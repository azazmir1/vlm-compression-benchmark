#!/usr/bin/env python3
"""
Unified VLM Profiling Pipeline

Profiles ALL models from the Excel tracking file — both FP16 baselines
and quantized variants (HQQ INT4, PyTorch INT8/INT4, GPTQ INT4).

For each model/method combination:
1. Loads the model (FP16 or quantized from HuggingFace)
2. Runs warmup inference
3. Profiles with:
   - TokenTimingProcessor: exact prefill/decode per-token timing
   - HookManager: component-level breakdown (attention, ffn, vision, etc.)
   - TegraStatsMonitor: power, temperature, RAM (Jetson-specific)
   - GPUProfiler: peak memory, GPU utilization
4. Saves granular JSON results to results/profiling/

Usage:
    # Profile a single model (FP16 baseline)
    python scripts/profile_all.py --model_id HuggingFaceTB/SmolVLM-256M-Instruct

    # Profile a quantized model from HuggingFace
    python scripts/profile_all.py --hf_repo Azaz666/SmolVLM-2.2B-HQQ-INT4

    # Profile all models from the Excel file
    python scripts/profile_all.py --all

    # Profile all with component-level hooks
    python scripts/profile_all.py --all --components

    # Profile specific family
    python scripts/profile_all.py --family smolvlm

    # Customize runs
    python scripts/profile_all.py --model_id HuggingFaceTB/SmolVLM-256M-Instruct \
        --num_samples 10 --num_warmup 3 --max_tokens 50 --components
"""

import argparse
import json
import logging
import sys
import time
import gc
import traceback
from pathlib import Path
from typing import Optional, Dict, List, Tuple

import torch
from PIL import Image
from tqdm import tqdm

# Project imports
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from models.model_loader import load_model, unload_model, detect_family
from profiling.gpu_profiler import GPUProfiler
from profiling.token_timer import TokenTimingProcessor
from profiling.tegrastats_monitor import TegraStatsMonitor
from profiling.detailed_metrics import DetailedMetrics, ComponentMetrics, CategoryTiming
from evaluation.run_baseline import (
    load_vqav2, run_inference,
    _vqa_multi_metric, _pope_accuracy,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

RESULTS_DIR = Path(__file__).resolve().parents[1] / "results" / "profiling"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)


# ── Model registry: maps Excel names to HuggingFace model IDs ──────────────

MODEL_REGISTRY = {
    # SmolVLM
    "SmolVLM-256M-Instruct": "HuggingFaceTB/SmolVLM-256M-Instruct",
    "SmolVLM-500M-Instruct": "HuggingFaceTB/SmolVLM-500M-Instruct",
    "SmolVLM-2.2B-Instruct": "HuggingFaceTB/SmolVLM-Instruct",
    # LFM2-VL
    "LFM2-VL-450M":  "LiquidAI/LFM2-VL-450M",
    "LFM2-VL-1.6B":  "LiquidAI/LFM2-VL-1.6B",
    "LFM2-VL-3B":    "LiquidAI/LFM2-VL-3B",
    # InternVL2.5
    "InternVL2.5-1B": "OpenGVLab/InternVL2_5-1B",
    "InternVL2.5-2B": "OpenGVLab/InternVL2_5-2B",
    "InternVL2.5-4B": "OpenGVLab/InternVL2_5-4B",
    "InternVL2.5-8B": "OpenGVLab/InternVL2_5-8B",
    # Qwen2.5-VL
    "Qwen2.5-VL-3B-Instruct": "Qwen/Qwen2.5-VL-3B-Instruct",
    "Qwen2.5-VL-7B-Instruct": "Qwen/Qwen2.5-VL-7B-Instruct",
    # Ovis2
    "Ovis2-1B": "AIDC-AI/Ovis2-1B",
    "Ovis2-2B": "AIDC-AI/Ovis2-2B",
    "Ovis2-4B": "AIDC-AI/Ovis2-4B",
    "Ovis2-8B": "AIDC-AI/Ovis2-8B",
    # Moondream
    "moondream2": "vikhyatk/moondream2",
    # Gemma 3
    "gemma-3-4b-it":  "google/gemma-3-4b-it",
    "gemma-3-12b-it": "google/gemma-3-12b-it",
    # FastVLM
    "FastVLM-0.5B": "apple/FastVLM-0.5B",
    "FastVLM-1.5B": "apple/FastVLM-1.5B",
}

# Maps (model_variant, method) -> HF quantized repo
# Only entries that have quantized models available
QUANTIZED_REGISTRY = {
    # SmolVLM-2.2B (HF repo uses "SmolVLM-Instruct" not "SmolVLM-2.2B")
    ("SmolVLM-2.2B-Instruct", "hqq_int4"):     "Azaz666/SmolVLM-Instruct-HQQ-INT4",
    ("SmolVLM-2.2B-Instruct", "pytorch_int8"):  "Azaz666/SmolVLM-Instruct-PYTORCH-INT8",
    ("SmolVLM-2.2B-Instruct", "pytorch_int4"):  "Azaz666/SmolVLM-Instruct-PYTORCH-INT4",
    ("SmolVLM-2.2B-Instruct", "gptq_int4"):     "Azaz666/SmolVLM-Instruct-GPTQ-INT4",
    # LFM2-VL-1.6B
    ("LFM2-VL-1.6B", "hqq_int4"):     "Azaz666/LFM2-VL-1.6B-HQQ-INT4",
    ("LFM2-VL-1.6B", "pytorch_int8"):  "Azaz666/LFM2-VL-1.6B-PYTORCH-INT8",
    ("LFM2-VL-1.6B", "pytorch_int4"):  "Azaz666/LFM2-VL-1.6B-PYTORCH-INT4",
    ("LFM2-VL-1.6B", "gptq_int4"):     "Azaz666/LFM2-VL-1.6B-GPTQ-INT4",
    # LFM2-VL-3B
    ("LFM2-VL-3B", "hqq_int4"):     "Azaz666/LFM2-VL-3B-HQQ-INT4",
    ("LFM2-VL-3B", "pytorch_int8"):  "Azaz666/LFM2-VL-3B-PYTORCH-INT8",
    ("LFM2-VL-3B", "pytorch_int4"):  "Azaz666/LFM2-VL-3B-PYTORCH-INT4",
    ("LFM2-VL-3B", "gptq_int4"):     "Azaz666/LFM2-VL-3B-GPTQ-INT4",
    # InternVL2.5-4B (HF repo uses underscore: InternVL2_5; only pytorch_int8/int4 uploaded)
    ("InternVL2.5-4B", "pytorch_int8"):  "Azaz666/InternVL2_5-4B-PYTORCH-INT8",
    ("InternVL2.5-4B", "pytorch_int4"):  "Azaz666/InternVL2_5-4B-PYTORCH-INT4",
    # Qwen2.5-VL-3B (HF repo uses full "Qwen2.5-VL-3B-Instruct-" prefix)
    ("Qwen2.5-VL-3B-Instruct", "hqq_int4"):     "Azaz666/Qwen2.5-VL-3B-Instruct-HQQ-INT4",
    ("Qwen2.5-VL-3B-Instruct", "pytorch_int8"):  "Azaz666/Qwen2.5-VL-3B-Instruct-PYTORCH-INT8",
    ("Qwen2.5-VL-3B-Instruct", "pytorch_int4"):  "Azaz666/Qwen2.5-VL-3B-Instruct-PYTORCH-INT4",
    ("Qwen2.5-VL-3B-Instruct", "gptq_int4"):     "Azaz666/Qwen2.5-VL-3B-Instruct-GPTQ-Int4",
    # Gemma-3-4B (only pytorch_int8/int4 uploaded)
    ("gemma-3-4b-it", "pytorch_int8"):  "Azaz666/gemma-3-4b-it-PYTORCH-INT8",
    ("gemma-3-4b-it", "pytorch_int4"):  "Azaz666/gemma-3-4b-it-PYTORCH-INT4",
    # Gemma-3-12B
    ("gemma-3-12b-it", "pytorch_int8"):  "Azaz666/gemma-3-12b-it-PYTORCH-INT8",
    ("gemma-3-12b-it", "pytorch_int4"):  "Azaz666/gemma-3-12b-it-PYTORCH-INT4",
}

FAMILY_MAP = {
    "SmolVLM": "smolvlm",
    "LFM2-VL": "lfm2vl",
    "InternVL2.5": "internvl25",
    "Qwen2.5-VL": "qwen25vl",
    "Ovis2": "ovis2",
    "Moondream": "moondream",
    "Gemma 3": "gemma3",
    "FastVLM": "fastvlm",
}


# ── Profiling functions ─────────────────────────────────────────────────────

def profile_single_sample(
    model, processor, sample: dict, family: str, device: str,
    max_new_tokens: int = 50,
    use_components: bool = False,
    hook_manager=None,
) -> Dict:
    """
    Profile a single VQA sample with granular timing.

    Returns dict with per-token timing, component breakdown, memory stats,
    and multi-metric accuracy scores.
    """
    image = sample["image"]
    question = sample["question"] + " Answer with a single word or short phrase."

    if not isinstance(image, Image.Image):
        image = Image.fromarray(image)
    image = image.convert("RGB")

    # ── 1. Preprocessing timing (CPU: tokenize + image transform) ─
    t_preprocess_start = time.perf_counter()
    inputs, gen_kwargs = _prepare_inputs(
        model, processor, image, question, family, device, max_new_tokens
    )
    t_preprocess_end = time.perf_counter()
    preprocess_ms = (t_preprocess_end - t_preprocess_start) * 1000

    # Count input tokens — handle both dict and BatchEncoding (UserDict)
    input_ids = None
    if hasattr(inputs, 'get'):
        input_ids = inputs.get("input_ids")
    elif isinstance(inputs, dict):
        input_ids = inputs.get("input_ids")
    num_input_tokens = input_ids.shape[1] if input_ids is not None else 0

    # ── 2. Memory before inference ───────────────────────────────
    if torch.cuda.is_available():
        torch.cuda.synchronize()
        torch.cuda.reset_peak_memory_stats()
        mem_before = torch.cuda.memory_allocated() / 1024**2
    else:
        mem_before = 0.0

    # ── 3. Generation with TokenTimingProcessor ──────────────────
    token_timer = TokenTimingProcessor()

    # Reset hooks if component profiling
    if use_components and hook_manager:
        hook_manager.reset()

    # Families that DON'T use model.generate() need wall-clock fallback
    # moondream: custom .answer_question() API
    # internvl25: uses model.chat() which wraps generate() internally
    needs_fallback = family in ("moondream", "internvl25")

    if not needs_fallback:
        # Create a token-aware LogitsProcessor that also updates hook token index
        class _TimerWithHookSync(TokenTimingProcessor):
            """TokenTimer that also updates hook_manager.current_token_idx."""
            def __init__(self, hook_mgr=None):
                super().__init__()
                self._hook_mgr = hook_mgr
                self._call_count = 0
            def __call__(self, input_ids, scores):
                # Update hook manager token index BEFORE recording event
                if self._hook_mgr:
                    self._hook_mgr.current_token_idx = self._call_count
                self._call_count += 1
                return super().__call__(input_ids, scores)

        token_timer = _TimerWithHookSync(hook_manager if use_components else None)

        existing_processors = gen_kwargs.pop("logits_processor", [])
        gen_kwargs["logits_processor"] = existing_processors + [token_timer]

        token_timer.record_start()
        with torch.no_grad():
            output_ids = model.generate(**inputs, **gen_kwargs)
        token_timer.record_end()
        token_timer.finalize()

        # Decode output
        t_decode_start = time.perf_counter()
        pred = _decode_output(output_ids, inputs, processor, family)
        t_decode_end = time.perf_counter()
        decode_ms = (t_decode_end - t_decode_start) * 1000
    else:
        # Fallback: wall-clock timing (moondream, internvl25)
        if torch.cuda.is_available():
            torch.cuda.synchronize()
            start_event = torch.cuda.Event(enable_timing=True)
            end_event = torch.cuda.Event(enable_timing=True)
            start_event.record()

        pred = run_inference(model, processor, sample, family, device, max_new_tokens)

        if torch.cuda.is_available():
            end_event.record()
            torch.cuda.synchronize()
            total_gen_ms = start_event.elapsed_time(end_event)
        else:
            total_gen_ms = 0.0

        token_timer.total_ms = total_gen_ms
        # Cannot determine exact token count without LogitsProcessor;
        # use tokenizer to count actual generated tokens
        try:
            if hasattr(processor, 'encode'):
                token_timer.num_tokens = len(processor.encode(pred)) if pred else 0
            elif hasattr(processor, 'tokenize'):
                token_timer.num_tokens = len(processor.tokenize(pred)) if pred else 0
            else:
                token_timer.num_tokens = len(pred.split()) if pred else 0
        except Exception:
            token_timer.num_tokens = len(pred.split()) if pred else 0
        token_timer._finalized = True
        decode_ms = 0.0

    # ── 4. Memory after inference ────────────────────────────────
    if torch.cuda.is_available():
        torch.cuda.synchronize()
        mem_peak = torch.cuda.max_memory_allocated() / 1024**2
        mem_after = torch.cuda.memory_allocated() / 1024**2
    else:
        mem_peak = mem_after = 0.0

    # ── 5. Component breakdown ───────────────────────────────────
    component_data = None
    if use_components and hook_manager:
        hook_manager.tracker.compute_timings()
        component_data = hook_manager.get_category_summary()

    # ── 6. Multi-metric accuracy ─────────────────────────────────
    gold_answers = sample.get("answers", [])
    if gold_answers and pred:
        metrics = _vqa_multi_metric(pred.strip(), gold_answers)
    else:
        metrics = {"exact_match": 0.0, "contains": 0.0, "token_f1": 0.0,
                   "bleu": 0.0, "rouge_l": 0.0}

    # ── 7. Assemble result ───────────────────────────────────────
    result = {
        "preprocessing_ms": round(preprocess_ms, 4),
        "num_input_tokens": num_input_tokens,
        "token_timing": token_timer.to_dict(),
        "decode_ms": round(decode_ms, 4),
        "total_ms": round(preprocess_ms + token_timer.total_ms + decode_ms, 4),
        "gpu_mem_before_mb": round(mem_before, 1),
        "gpu_mem_peak_mb": round(mem_peak, 1),
        "gpu_mem_after_mb": round(mem_after, 1),
        "generated_text": pred.strip() if pred else "",
        "accuracy": metrics,
    }

    if component_data:
        result["components"] = {
            cat: {
                "total_ms": round(stats["total_ms"], 2),
                "percentage": round(stats["percentage"], 1),
                "count": stats["count"],
            }
            for cat, stats in component_data.items()
        }

    return result


def _prepare_inputs(model, processor, image, question, family, device, max_new_tokens):
    """
    Prepare model inputs for generation. Returns (inputs_dict, gen_kwargs).
    Mirrors the logic in run_inference() but returns inputs instead of running.
    """
    gen_kwargs = {"max_new_tokens": max_new_tokens, "do_sample": False}

    if family == "florence2":
        inputs = processor(
            text=f"<VQA> {question}", images=image, return_tensors="pt"
        ).to(device)
        inputs["pixel_values"] = inputs["pixel_values"].to(dtype=torch.float16)
        gen_kwargs = {"max_new_tokens": max_new_tokens, "num_beams": 1}
        return inputs, gen_kwargs

    elif family == "smolvlm":
        messages = [{"role": "user", "content": [{"type": "image"}, {"type": "text", "text": question}]}]
        prompt = processor.apply_chat_template(messages, add_generation_prompt=True)
        inputs = processor(text=prompt, images=[image], return_tensors="pt").to(device)
        return inputs, gen_kwargs

    elif family == "fastvlm":
        IMAGE_TOKEN_INDEX = -200
        messages = [{"role": "user", "content": f"<image>\n{question}"}]
        rendered = processor.apply_chat_template(messages, add_generation_prompt=True, tokenize=False)
        pre, post = rendered.split("<image>", 1)
        pre_ids = processor.tokenizer(pre, return_tensors="pt", add_special_tokens=False).input_ids
        post_ids = processor.tokenizer(post, return_tensors="pt", add_special_tokens=False).input_ids
        img_tok = torch.tensor([[IMAGE_TOKEN_INDEX]], dtype=pre_ids.dtype)
        input_ids = torch.cat([pre_ids, img_tok, post_ids], dim=1).to(device)
        attention_mask = torch.ones_like(input_ids, device=device)
        pixel_values = model.get_vision_tower().image_processor(
            images=image, return_tensors="pt"
        )["pixel_values"].to(device, dtype=next(model.parameters()).dtype)
        inputs = {"inputs": input_ids, "attention_mask": attention_mask, "images": pixel_values}
        return inputs, gen_kwargs

    elif family == "lfm2vl":
        messages = [{"role": "user", "content": [
            {"type": "image", "image": image}, {"type": "text", "text": question}
        ]}]
        text = processor.apply_chat_template(messages, add_generation_prompt=True)
        inputs = processor(text=text, images=[image], return_tensors="pt").to(device)
        return inputs, gen_kwargs

    elif family == "qwen25vl":
        messages = [{"role": "user", "content": [
            {"type": "image", "image": image}, {"type": "text", "text": question}
        ]}]
        text = processor.apply_chat_template(messages, add_generation_prompt=True)
        inputs = processor(text=[text], images=[image], return_tensors="pt").to(device)
        return inputs, gen_kwargs

    elif family == "gemma3":
        messages = [{"role": "user", "content": [
            {"type": "image", "image": image}, {"type": "text", "text": question}
        ]}]
        inputs = processor.apply_chat_template(
            messages, add_generation_prompt=True,
            tokenize=True, return_dict=True, return_tensors="pt"
        ).to(device)
        return inputs, gen_kwargs

    elif family == "ovis2":
        text_tokenizer = processor
        ovis_query = f"<image>\n{question}"
        prompt, input_ids, pixel_values = model.preprocess_inputs(ovis_query, [image])
        attention_mask = torch.ne(input_ids, text_tokenizer.pad_token_id)
        input_ids = input_ids.unsqueeze(0).to(device)
        attention_mask = attention_mask.unsqueeze(0).to(device)
        pixel_values = [pixel_values.to(dtype=next(model.parameters()).dtype, device=device)]
        gen_kwargs.update({
            "eos_token_id": model.generation_config.eos_token_id,
            "pad_token_id": text_tokenizer.pad_token_id,
        })
        inputs = {"input_ids": input_ids, "pixel_values": pixel_values, "attention_mask": attention_mask}
        return inputs, gen_kwargs

    elif family == "internvl25":
        # InternVL uses model.chat() which doesn't support LogitsProcessor
        # Fall back to preparing inputs manually
        import torchvision.transforms as T
        transform = T.Compose([
            T.Resize((448, 448)), T.ToTensor(),
            T.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
        ])
        pixel_values = transform(image).unsqueeze(0).to(device=device, dtype=torch.float16)
        tokenizer = processor
        question_fmt = f"<image>\n{question}"
        # InternVL's chat() method wraps generate() internally
        # We'll use model.chat() in the fallback path
        inputs = {"pixel_values": pixel_values, "_question": question_fmt, "_tokenizer": tokenizer}
        return inputs, gen_kwargs

    elif family == "moondream":
        # moondream uses custom API, not .generate()
        inputs = {"_sample": {"image": image, "question": question}}
        return inputs, gen_kwargs

    else:
        raise ValueError(f"Unknown family '{family}'")


def _decode_output(output_ids, inputs, processor, family):
    """Decode generated token IDs to text."""
    if family in ("smolvlm", "lfm2vl", "qwen25vl"):
        input_len = inputs["input_ids"].shape[1]
        return processor.batch_decode(
            output_ids[:, input_len:], skip_special_tokens=True
        )[0]
    elif family == "gemma3":
        input_len = inputs["input_ids"].shape[1]
        return processor.decode(output_ids[0][input_len:], skip_special_tokens=True)
    elif family == "fastvlm":
        pred = processor.batch_decode(output_ids, skip_special_tokens=True)[0]
        return pred.split("\n")[0].strip()
    elif family == "florence2":
        return processor.batch_decode(output_ids, skip_special_tokens=True)[0]
    elif family == "ovis2":
        text_tokenizer = processor
        return text_tokenizer.decode(output_ids[0], skip_special_tokens=True)
    else:
        return processor.batch_decode(output_ids, skip_special_tokens=True)[0]


# ── Main profiling loop ─────────────────────────────────────────────────────

def profile_model(
    model, processor, family: str, device: str,
    model_id: str, method: str,
    samples: list,
    num_warmup: int = 3,
    max_tokens: int = 50,
    use_components: bool = False,
) -> Dict:
    """
    Profile a model on VQA samples with full instrumentation.
    """
    # Setup component hooks
    hook_manager = None
    if use_components:
        from profiling.hooks import HookManager, TimingTracker, ModuleCategorizer
        categorizer = ModuleCategorizer()
        tracker = TimingTracker()
        hook_manager = HookManager(categorizer, tracker)
        hook_manager.register_hooks(model)

    # Warmup
    if num_warmup > 0:
        logger.info(f"Warming up ({num_warmup} runs)...")
        if use_components and hook_manager:
            hook_manager.disable()
        for i in range(num_warmup):
            try:
                run_inference(model, processor, samples[i % len(samples)], family, device, max_tokens)
            except Exception:
                pass
        if use_components and hook_manager:
            hook_manager.enable()
        # Clear GPU cache after warmup
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

    # Start tegrastats monitoring
    tegra = TegraStatsMonitor(interval_ms=200)
    tegra.start()

    # Profile each sample
    sample_results = []
    gpu_profiler = GPUProfiler(device_index=0, poll_interval_ms=100)

    logger.info(f"Profiling {len(samples)} samples...")
    with gpu_profiler:
        for sample in tqdm(samples, desc=f"{method}"):
            try:
                result = profile_single_sample(
                    model, processor, sample, family, device,
                    max_new_tokens=max_tokens,
                    use_components=use_components,
                    hook_manager=hook_manager,
                )
                sample_results.append(result)
            except Exception as e:
                logger.warning(f"Sample failed: {e}")
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()

    tegra.stop()
    gpu_stats = gpu_profiler.stats()
    tegra_stats = tegra.stats()

    # Cleanup hooks
    if hook_manager:
        hook_manager.remove_hooks()

    # ── Aggregate results ────────────────────────────────────────
    if not sample_results:
        logger.error("All samples failed!")
        return {"error": "all_samples_failed"}

    n = len(sample_results)

    # Per-token timing aggregation
    all_prefill = [r["token_timing"]["prefill_ms"] for r in sample_results if r["token_timing"].get("prefill_ms")]
    all_decode = [r["token_timing"]["decode_ms"] for r in sample_results if r["token_timing"].get("decode_ms")]
    all_total = [r["token_timing"]["total_ms"] for r in sample_results if r["token_timing"].get("total_ms")]
    all_tokens = [r["token_timing"]["num_tokens"] for r in sample_results if r["token_timing"].get("num_tokens")]

    avg_prefill = sum(all_prefill) / len(all_prefill) if all_prefill else 0.0
    avg_decode = sum(all_decode) / len(all_decode) if all_decode else 0.0
    avg_total_gen = sum(all_total) / len(all_total) if all_total else 0.0
    avg_tokens = sum(all_tokens) / len(all_tokens) if all_tokens else 0.0

    # Decode throughput
    decode_throughputs = []
    for r in sample_results:
        tt = r["token_timing"]
        n_decode = tt.get("num_tokens", 0) - 1
        decode_ms = tt.get("decode_ms", 0)
        if n_decode > 0 and decode_ms > 0:
            decode_throughputs.append(n_decode / (decode_ms / 1000.0))

    # Component aggregation
    component_agg = {}
    if use_components:
        for r in sample_results:
            for cat, stats in r.get("components", {}).items():
                if cat not in component_agg:
                    component_agg[cat] = {"total_ms": 0, "count": 0, "samples": 0}
                component_agg[cat]["total_ms"] += stats["total_ms"]
                component_agg[cat]["count"] += stats["count"]
                component_agg[cat]["samples"] += 1
        # Average and compute percentages
        total_comp_ms = sum(c["total_ms"] for c in component_agg.values())
        for cat in component_agg:
            c = component_agg[cat]
            c["avg_ms"] = round(c["total_ms"] / c["samples"], 2)
            c["percentage"] = round((c["total_ms"] / total_comp_ms) * 100, 1) if total_comp_ms > 0 else 0.0
            c["total_ms"] = round(c["total_ms"], 2)

    # ── Accuracy aggregation ─────────────────────────────────────
    metric_names = ["exact_match", "contains", "token_f1", "bleu", "rouge_l"]
    accuracy_agg = {}
    for name in metric_names:
        vals = [r["accuracy"][name] for r in sample_results if "accuracy" in r]
        accuracy_agg[name] = round(sum(vals) / len(vals), 4) if vals else 0.0

    # Build final result
    result = {
        "model_id": model_id,
        "method": method,
        "family": family,
        "num_samples": n,
        "num_warmup": num_warmup,
        "max_tokens": max_tokens,

        # Multi-metric accuracy
        "accuracy": accuracy_agg,

        # Timing summary
        "timing": {
            "avg_preprocessing_ms": round(sum(r["preprocessing_ms"] for r in sample_results) / n, 2),
            "avg_prefill_ms": round(avg_prefill, 2),
            "avg_decode_ms": round(avg_decode, 2),
            "avg_total_generation_ms": round(avg_total_gen, 2),
            "avg_decode_detokenize_ms": round(sum(r["decode_ms"] for r in sample_results) / n, 2),
            "avg_total_ms": round(sum(r["total_ms"] for r in sample_results) / n, 2),
            "avg_tokens_generated": round(avg_tokens, 1),
            "avg_input_tokens": round(sum(r["num_input_tokens"] for r in sample_results) / n, 1),
        },

        # Throughput
        "throughput": {
            "avg_tok_s": round(avg_tokens / (avg_total_gen / 1000.0), 2) if avg_total_gen > 0 else 0.0,
            "avg_decode_tok_s": round(sum(decode_throughputs) / len(decode_throughputs), 2) if decode_throughputs else 0.0,
            "samples_per_s": round(n / gpu_stats.wall_time_s, 3) if gpu_stats.wall_time_s > 0 else 0.0,
        },

        # Memory
        "memory": {
            "gpu_peak_mb": round(gpu_stats.peak_memory_mb, 1),
            "gpu_avg_mb": round(gpu_stats.avg_memory_mb, 1),
            "avg_inference_peak_mb": round(sum(r["gpu_mem_peak_mb"] for r in sample_results) / n, 1),
        },

        # GPU utilization (from pynvml)
        "gpu_util": {
            "avg_pct": round(gpu_stats.avg_gpu_util_pct, 1),
            "peak_pct": round(gpu_stats.peak_gpu_util_pct, 1),
        },

        # Tegrastats (Jetson — power, temp, RAM, GPU util)
        "tegrastats": tegra_stats.to_dict(),

        # Per-sample detailed results (first 3)
        "sample_details": sample_results[:3],
    }

    # Component breakdown
    if component_agg:
        result["components"] = component_agg

    return result


def load_model_for_profiling(
    model_id: str, method: str = "fp16", hf_repo: Optional[str] = None
) -> Tuple:
    """
    Load a model for profiling. Handles both FP16 baseline and quantized variants.

    For FP16: uses standard model_loader.
    For quantized: uses benchmark_compressed_hf.py loading path.
    """
    if method == "fp16":
        model, processor, meta = load_model(model_id)
        family = meta.family
        try:
            device = str(next(model.parameters()).device)
        except StopIteration:
            device = "cuda:0"
        return model, processor, family, device

    # Quantized model — delegate to benchmark_compressed_hf loading
    if hf_repo is None:
        raise ValueError(f"hf_repo required for method={method}")

    # Import the loading functions from benchmark_compressed_hf
    from profiling.benchmark_compressed_hf import (
        detect_format,
        load_hqq_int4, load_pytorch_int8, load_pytorch_int4, load_gptq_int4
    )

    family = detect_family(model_id)
    fmt, cfg = detect_format(hf_repo)
    logger.info(f"Detected format: {fmt} for {hf_repo}")

    base_model = cfg.get("base_model") or model_id

    # Each loader takes (hf_repo, base_model, family) and returns (model, processor)
    loaders = {
        "hqq_int4": load_hqq_int4,
        "pytorch_int8": load_pytorch_int8,
        "pytorch_int4": load_pytorch_int4,
        "gptq_int4": load_gptq_int4,
    }
    loader = loaders.get(fmt)
    if loader is None:
        raise ValueError(f"No loader for format {fmt}")

    model, processor = loader(hf_repo, base_model, family)

    device = "cuda:0"
    return model, processor, family, device


# ── Excel integration ────────────────────────────────────────────────────────

def get_models_from_excel(excel_path: str = None) -> List[Dict]:
    """
    Read model list from Excel file and return list of profiling targets.

    Returns list of dicts:
        {"variant": str, "family": str, "model_id": str,
         "method": str, "hf_repo": str or None, "status": str}
    """
    if excel_path is None:
        excel_path = str(Path(__file__).resolve().parents[1] / "VLM_Model_Families_Jetson_Status.xlsx")

    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    targets = []
    methods_cols = {
        4: "hqq_int4",
        5: "pytorch_int8",
        6: "pytorch_int4",
        7: "gptq_int4",
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        family_name = row[0].value
        variant = row[1].value
        status = row[3].value

        if not variant or not family_name:
            continue

        model_id = MODEL_REGISTRY.get(variant)
        if not model_id:
            continue

        family_code = FAMILY_MAP.get(family_name, "")

        # FP16 baseline (only if PASS or MEM_CRITICAL — model can load)
        if status in ("PASS",):
            targets.append({
                "variant": variant,
                "family": family_code,
                "model_id": model_id,
                "method": "fp16",
                "hf_repo": None,
                "status": status,
            })

        # Quantized variants (only if accuracy value exists in Excel)
        for col_idx, method in methods_cols.items():
            val = row[col_idx].value
            if val and val != "--" and isinstance(val, (int, float)):
                hf_repo = QUANTIZED_REGISTRY.get((variant, method))
                if hf_repo:
                    targets.append({
                        "variant": variant,
                        "family": family_code,
                        "model_id": model_id,
                        "method": method,
                        "hf_repo": hf_repo,
                        "status": status,
                    })

    return targets


def filter_best_quantized(targets: List[Dict]) -> List[Dict]:
    """
    For OOM_LOAD models with multiple quantization methods, keep only
    the method with the highest accuracy in the Excel file.
    FP16 (PASS) models are always kept.
    """
    # Group quantized targets by variant
    from collections import defaultdict
    variant_methods = defaultdict(list)
    fp16_targets = []

    for t in targets:
        if t["method"] == "fp16":
            fp16_targets.append(t)
        else:
            variant_methods[t["variant"]].append(t)

    # For each variant, find the method with highest accuracy from Excel
    import openpyxl
    excel_path = str(Path(__file__).resolve().parents[1] / "VLM_Model_Families_Jetson_Status.xlsx")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    methods_cols = {4: "hqq_int4", 5: "pytorch_int8", 6: "pytorch_int4", 7: "gptq_int4"}

    variant_best = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        variant = row[1].value
        if not variant:
            continue
        best_acc, best_method = -1, None
        for col_idx, method in methods_cols.items():
            val = row[col_idx].value
            if val and val != "--" and isinstance(val, (int, float)) and val > best_acc:
                best_acc = val
                best_method = method
        if best_method:
            variant_best[variant] = best_method

    # Filter: keep only the best method per variant (that has an available repo)
    best_targets = list(fp16_targets)
    for variant, methods in variant_methods.items():
        best = variant_best.get(variant)
        # Try exact best method first
        chosen = None
        for t in methods:
            if t["method"] == best:
                chosen = t
                break
        # If best method doesn't have a repo, pick the available method with highest accuracy
        if chosen is None and methods:
            # Re-read Excel to get accuracy for each available method
            avail_methods = {t["method"] for t in methods}
            best_avail_acc, best_avail = -1, None
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
                if row[1].value != variant:
                    continue
                for col_idx, method in methods_cols.items():
                    if method not in avail_methods:
                        continue
                    val = row[col_idx].value
                    if val and val != "--" and isinstance(val, (int, float)) and val > best_avail_acc:
                        best_avail_acc = val
                        best_avail = method
            for t in methods:
                if t["method"] == best_avail:
                    chosen = t
                    break
        if chosen:
            best_targets.append(chosen)

    return best_targets


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Unified VLM Profiling Pipeline")

    # Model selection (mutually exclusive modes)
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--model_id", help="HuggingFace model ID for FP16 profiling")
    group.add_argument("--hf_repo", help="HuggingFace repo for quantized model profiling")
    group.add_argument("--all", action="store_true", help="Profile all models from Excel")
    group.add_argument("--family", help="Profile all models of a specific family")

    # Profiling options
    parser.add_argument("--method", default="fp16",
                        choices=["fp16", "hqq_int4", "pytorch_int8", "pytorch_int4", "gptq_int4"],
                        help="Quantization method (for --model_id mode)")
    parser.add_argument("--base_model", help="Base model ID (for --hf_repo mode)")
    parser.add_argument("--num_samples", type=int, default=10, help="Number of VQA samples")
    parser.add_argument("--num_warmup", type=int, default=3, help="Number of warmup runs")
    parser.add_argument("--max_tokens", type=int, default=50, help="Max tokens to generate")
    parser.add_argument("--components", action="store_true", help="Enable component-level hook profiling")
    parser.add_argument("--best-only", action="store_true",
                        help="For OOM models, only profile the best quantization method")
    parser.add_argument("--force", action="store_true", help="Overwrite existing results")
    parser.add_argument("--excel", help="Path to Excel tracking file")

    args = parser.parse_args()

    # Load VQA samples
    samples = load_vqav2(n_samples=args.num_samples)

    if args.all or args.family:
        # Batch mode: profile all models from Excel
        targets = get_models_from_excel(args.excel)
        if args.best_only:
            targets = filter_best_quantized(targets)
        if args.family:
            family_code = FAMILY_MAP.get(args.family, args.family)
            targets = [t for t in targets if t["family"] == family_code]

        logger.info(f"Found {len(targets)} profiling targets")

        for i, target in enumerate(targets):
            safe_name = f"{target['variant']}__{target['method']}"
            out_path = RESULTS_DIR / f"{safe_name}.json"

            if out_path.exists() and not args.force:
                logger.info(f"[{i+1}/{len(targets)}] Skip {safe_name} (exists)")
                continue

            logger.info(f"\n[{i+1}/{len(targets)}] Profiling {safe_name}")
            try:
                model, processor, family, device = load_model_for_profiling(
                    target["model_id"], target["method"], target["hf_repo"]
                )

                result = profile_model(
                    model, processor, family, device,
                    target["model_id"], target["method"],
                    samples,
                    num_warmup=args.num_warmup,
                    max_tokens=args.max_tokens,
                    use_components=args.components,
                )

                with open(out_path, "w") as f:
                    json.dump(result, f, indent=2)
                logger.info(f"Saved to {out_path}")

                # Print summary
                if "timing" in result:
                    t = result["timing"]
                    tp = result["throughput"]
                    acc = result.get("accuracy", {})
                    logger.info(
                        f"  Prefill={t['avg_prefill_ms']:.1f}ms  "
                        f"Decode={t['avg_decode_ms']:.1f}ms  "
                        f"Total={t['avg_total_ms']:.1f}ms  "
                        f"Throughput={tp['avg_decode_tok_s']:.1f} tok/s  "
                        f"Mem={result['memory']['gpu_peak_mb']:.0f}MB  "
                        f"EM={acc.get('exact_match', 0):.3f}"
                    )

                unload_model(model)

            except Exception as e:
                logger.error(f"  FAILED: {e}")
                traceback.print_exc()

            gc.collect()
            if torch.cuda.is_available():
                torch.cuda.empty_cache()

    else:
        # Single model mode
        if args.hf_repo:
            if not args.base_model:
                logger.error("--base_model required with --hf_repo")
                sys.exit(1)
            model_id = args.base_model
            method = "auto"  # will be detected
            hf_repo = args.hf_repo
        else:
            model_id = args.model_id
            method = args.method
            hf_repo = None
            if method != "fp16":
                # Look up in registry
                family = detect_family(model_id)
                variant = next((k for k, v in MODEL_REGISTRY.items() if v == model_id), None)
                if variant:
                    hf_repo = QUANTIZED_REGISTRY.get((variant, method))

        model, processor, family, device = load_model_for_profiling(
            model_id, method, hf_repo
        )

        result = profile_model(
            model, processor, family, device,
            model_id, method,
            samples,
            num_warmup=args.num_warmup,
            max_tokens=args.max_tokens,
            use_components=args.components,
        )

        safe_name = model_id.replace("/", "__") + f"__{method}"
        out_path = RESULTS_DIR / f"{safe_name}.json"
        with open(out_path, "w") as f:
            json.dump(result, f, indent=2)
        logger.info(f"Saved to {out_path}")

        # Print detailed summary
        if "timing" in result:
            t = result["timing"]
            tp = result["throughput"]
            m = result["memory"]
            print("\n" + "=" * 70)
            print(f"Profiling Results: {model_id} [{method}]")
            print("=" * 70)
            print(f"  Samples:             {result['num_samples']}")
            print(f"  Avg tokens:          {t['avg_tokens_generated']:.0f}")
            print(f"  Preprocessing:       {t['avg_preprocessing_ms']:8.2f} ms")
            print(f"  Prefill (1st tok):   {t['avg_prefill_ms']:8.2f} ms")
            print(f"  Decode:              {t['avg_decode_ms']:8.2f} ms")
            print(f"  Total generation:    {t['avg_total_generation_ms']:8.2f} ms")
            print(f"  Total (incl prep):   {t['avg_total_ms']:8.2f} ms")
            print(f"  Throughput:          {tp['avg_tok_s']:8.2f} tok/s")
            print(f"  Decode throughput:   {tp['avg_decode_tok_s']:8.2f} tok/s")
            print(f"  GPU peak memory:     {m['gpu_peak_mb']:8.1f} MB")

            if result.get("tegrastats", {}).get("num_samples", 0) > 0:
                ts = result["tegrastats"]
                print(f"  Avg power:           {ts['avg_power_w']:8.1f} W")
                print(f"  Avg GPU temp:        {ts['avg_gpu_temp_c']:8.1f} C")

            if result.get("components"):
                print("\n  Component Breakdown:")
                sorted_comps = sorted(
                    result["components"].items(),
                    key=lambda x: x[1]["total_ms"],
                    reverse=True
                )
                for cat, stats in sorted_comps:
                    print(f"    {cat:<25} {stats['avg_ms']:8.2f} ms/sample  ({stats['percentage']:.1f}%)")

            print("=" * 70)

        unload_model(model)


if __name__ == "__main__":
    main()
