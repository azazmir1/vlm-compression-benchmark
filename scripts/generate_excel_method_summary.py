#!/usr/bin/env python3
"""Generate Excel file summarizing all methods: success rates, per-model breakdown, and tier analysis."""

import json
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS_DIR = os.path.join(os.path.dirname(__file__), "..", "results")

ALL_MODELS = [
    "HuggingFaceTB/SmolVLM-256M-Instruct",
    "LiquidAI/LFM2-VL-450M",
    "HuggingFaceTB/SmolVLM-500M-Instruct",
    "OpenGVLab/InternVL2_5-1B",
    "AIDC-AI/Ovis2-1B",
    "LiquidAI/LFM2-VL-1.6B",
    "OpenGVLab/InternVL2_5-2B",
    "AIDC-AI/Ovis2-2B",
    "vikhyatk/moondream2",
    "HuggingFaceTB/SmolVLM-Instruct",
    "LiquidAI/LFM2-VL-3B",
    "Qwen/Qwen2.5-VL-3B-Instruct",
    "OpenGVLab/InternVL2_5-4B",
    "AIDC-AI/Ovis2-4B",
    "google/gemma-3-4b-it",
    "Qwen/Qwen2.5-VL-7B-Instruct",
    "OpenGVLab/InternVL2_5-8B",
    "AIDC-AI/Ovis2-8B",
    "google/gemma-3-12b-it",
]

# Method definitions: (display_name, category, subdir, filter_func, config_desc)
METHOD_DEFS = [
    ("Baseline (FP16)", "—", "baseline", lambda d: True, "FP16 (no compression)"),
    ("PTQ (BnB) INT4", "Quantization", "ptq", lambda d: d.get("quant") == "int4", "INT4 BitsAndBytes"),
    ("PTQ (BnB) INT8", "Quantization", "ptq", lambda d: d.get("quant") == "int8", "INT8 BitsAndBytes"),
    ("Magnitude Pruning 20%", "Pruning", "pruning", lambda d: abs(d.get("target_sparsity", 0) - 0.2) < 0.05, "L1 Unstructured 20%"),
    ("Magnitude Pruning 40%", "Pruning", "pruning", lambda d: abs(d.get("target_sparsity", 0) - 0.4) < 0.05, "L1 Unstructured 40%"),
    ("Wanda 20%", "Pruning", "wanda", lambda d: abs(d.get("target_sparsity", 0) - 0.2) < 0.05, "Wanda 20% sparsity"),
    ("Wanda 40%", "Pruning", "wanda", lambda d: abs(d.get("target_sparsity", 0) - 0.4) < 0.05, "Wanda 40% sparsity"),
    ("AWQ (Simulated)", "Quantization", "awq_gptq", lambda d: d.get("method") == "awq", "Sim. INT4 per-channel"),
    ("GPTQ (Simulated)", "Quantization", "awq_gptq", lambda d: d.get("method") == "gptq", "Sim. INT4 per-channel"),
    ("SparseGPT 50%", "Pruning", "sparsegpt", lambda d: True, "50% sparsity, Hessian-based"),
    ("AWP (Prune+Quant)", "Combined", "awp", lambda d: True, "Wanda 50% + Sim. INT4"),
    ("PACT (Prune+Merge)", "Token Compression", "pact", lambda d: True, "Prune 30% + Merge 20%"),
    ("SVD-LLM", "Low-Rank", "svd_llm", lambda d: True, "MLP SVD energy=0.95"),
    ("PALU (KV Compression)", "Low-Rank", "palu", lambda d: True, "KV SVD energy=0.95"),
    ("CASP (Mixed Quant)", "Mixed Precision", "casp_slim", lambda d: d.get("method") == "casp", "8/4-bit + LowRank QK"),
    ("SLIM (SVD+Prune+Quant)", "Combined", "casp_slim", lambda d: d.get("method") == "slim", "SVD r=0.3 + Prune 50% + Quant"),
]

def load_results():
    """Load all results and organize by method and model."""
    # method_name -> {model_id -> result_data}
    method_results = {m[0]: {} for m in METHOD_DEFS}

    for method_name, category, subdir, filter_fn, config in METHOD_DEFS:
        dir_path = os.path.join(RESULTS_DIR, subdir)
        for json_path in sorted(glob.glob(os.path.join(dir_path, "*.json"))):
            if "gptq_comparison" in json_path:
                continue
            try:
                with open(json_path) as f:
                    data = json.load(f)
            except (json.JSONDecodeError, IOError):
                continue

            if not filter_fn(data):
                continue

            model_id = data.get("model_id", "")
            bench = data.get("benchmarks", {}).get("vqav2", {})
            acc = bench.get("accuracy")
            all_failed = bench.get("all_failed", False)

            success = (not all_failed) and (acc is not None) and (acc > 0)

            method_results[method_name][model_id] = {
                "success": success,
                "accuracy": acc if success else None,
                "latency": bench.get("avg_latency_s") if success else None,
                "peak_memory": bench.get("peak_memory_mb") if success else None,
                "throughput": bench.get("throughput_sps") if success else None,
                "power": bench.get("avg_power_w") if success else None,
                "data": data,
            }

    return method_results


def create_excel(method_results, output_path):
    wb = Workbook()
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold_font = Font(bold=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    def style_header(ws, cols):
        for ci, name in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=name)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
            c.border = thin_border

    # ═══════════════════════════════════════════════════════════
    # Sheet 1: Method Success Summary
    # ═══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Method Success Summary"

    cols1 = ["#", "Method", "Category", "Configuration", "Succeeded", "Total", "Success Rate (%)", "Tier",
             "Avg Accuracy (successful)", "Failed Models"]
    style_header(ws1, cols1)

    for row_idx, (method_name, category, subdir, filter_fn, config) in enumerate(METHOD_DEFS, 2):
        results = method_results[method_name]
        n_success = sum(1 for v in results.values() if v["success"])
        n_total = 19  # all models attempted
        rate = (n_success / n_total) * 100

        # Tier
        if rate == 100:
            tier = "100% Success"
        elif rate >= 70:
            tier = "High (>=70%)"
        elif rate >= 45:
            tier = "Moderate (45-69%)"
        else:
            tier = "Low (<45%)"

        # Avg accuracy of successful runs
        accs = [v["accuracy"] for v in results.values() if v["success"] and v["accuracy"] is not None]
        avg_acc = sum(accs) / len(accs) if accs else None

        # Failed models
        successful_models = {m for m, v in results.items() if v["success"]}
        failed = sorted(set(ALL_MODELS) - successful_models)
        failed_str = ", ".join([m.split("/")[-1] for m in failed]) if failed else "None"

        ws1.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
        ws1.cell(row=row_idx, column=2, value=method_name).border = thin_border
        ws1.cell(row=row_idx, column=3, value=category).border = thin_border
        ws1.cell(row=row_idx, column=4, value=config).border = thin_border

        c_succ = ws1.cell(row=row_idx, column=5, value=n_success)
        c_succ.border = thin_border
        c_succ.font = bold_font

        ws1.cell(row=row_idx, column=6, value=n_total).border = thin_border

        c_rate = ws1.cell(row=row_idx, column=7, value=round(rate, 1))
        c_rate.border = thin_border
        c_rate.number_format = '0.0'
        c_rate.font = bold_font
        if rate == 100:
            c_rate.fill = green_fill
        elif rate >= 70:
            c_rate.fill = PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid")
        elif rate >= 45:
            c_rate.fill = yellow_fill
        else:
            c_rate.fill = red_fill

        c_tier = ws1.cell(row=row_idx, column=8, value=tier)
        c_tier.border = thin_border
        if tier == "100% Success":
            c_tier.fill = green_fill
        elif "High" in tier:
            c_tier.fill = PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid")
        elif "Moderate" in tier:
            c_tier.fill = yellow_fill
        else:
            c_tier.fill = red_fill

        c_acc = ws1.cell(row=row_idx, column=9, value=avg_acc)
        c_acc.border = thin_border
        if avg_acc is not None:
            c_acc.number_format = '0.0000'

        ws1.cell(row=row_idx, column=10, value=failed_str).border = thin_border

    # Totals row
    total_row = len(METHOD_DEFS) + 2
    ws1.cell(row=total_row, column=2, value="TOTAL").font = bold_font
    total_s = sum(sum(1 for v in method_results[m[0]].values() if v["success"]) for m in METHOD_DEFS)
    total_t = 19 * len(METHOD_DEFS)
    ws1.cell(row=total_row, column=5, value=total_s).font = bold_font
    ws1.cell(row=total_row, column=6, value=total_t).font = bold_font
    c = ws1.cell(row=total_row, column=7, value=round((total_s / total_t) * 100, 1))
    c.font = bold_font
    c.number_format = '0.0'

    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 30
    ws1.column_dimensions["E"].width = 12
    ws1.column_dimensions["F"].width = 8
    ws1.column_dimensions["G"].width = 16
    ws1.column_dimensions["H"].width = 18
    ws1.column_dimensions["I"].width = 22
    ws1.column_dimensions["J"].width = 80
    ws1.auto_filter.ref = f"A1:J{len(METHOD_DEFS)+1}"
    ws1.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════════
    # Sheet 2: Per-Model × Method Grid (Success/Fail + Accuracy)
    # ═══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Model × Method Grid")

    method_names = [m[0] for m in METHOD_DEFS]
    cols2 = ["Model"] + method_names
    style_header(ws2, cols2)

    # Get baseline accuracies
    baseline_accs = {}
    for model_id in ALL_MODELS:
        r = method_results["Baseline (FP16)"].get(model_id, {})
        if r.get("success"):
            baseline_accs[model_id] = r["accuracy"]

    for row_idx, model_id in enumerate(ALL_MODELS, 2):
        short = model_id.split("/")[-1]
        ws2.cell(row=row_idx, column=1, value=short).border = thin_border

        base_acc = baseline_accs.get(model_id)

        for col_idx, method_name in enumerate(method_names, 2):
            r = method_results[method_name].get(model_id, {})
            if r.get("success"):
                acc = r["accuracy"]
                cell = ws2.cell(row=row_idx, column=col_idx, value=acc)
                cell.number_format = '0.0000'
                # Color by accuracy vs baseline
                if base_acc and acc is not None:
                    if acc >= base_acc:
                        cell.fill = green_fill
                    elif acc >= base_acc * 0.9:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill
                else:
                    cell.fill = green_fill
            elif model_id in [m for m, v in method_results[method_name].items()]:
                # Was attempted but failed
                cell = ws2.cell(row=row_idx, column=col_idx, value="FAILED")
                cell.fill = red_fill
                cell.font = Font(color="9C0006", bold=True, size=9)
            else:
                # Not attempted (method didn't produce a file for this model)
                cell = ws2.cell(row=row_idx, column=col_idx, value="N/A")
                cell.fill = gray_fill
                cell.font = Font(color="808080", size=9)
            cell.border = thin_border

    # Success count row at bottom
    count_row = len(ALL_MODELS) + 2
    ws2.cell(row=count_row, column=1, value="SUCCEEDED").font = bold_font
    for col_idx, method_name in enumerate(method_names, 2):
        n = sum(1 for v in method_results[method_name].values() if v["success"])
        c = ws2.cell(row=count_row, column=col_idx, value=f"{n}/19")
        c.font = bold_font
        c.alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 30
    for ci in range(2, len(method_names) + 2):
        ws2.column_dimensions[get_column_letter(ci)].width = 16
    ws2.freeze_panes = "B2"

    # ═══════════════════════════════════════════════════════════
    # Sheet 3: Per-Model × Method Grid (Latency)
    # ═══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Latency Grid")

    cols3 = ["Model"] + method_names
    style_header(ws3, cols3)

    for row_idx, model_id in enumerate(ALL_MODELS, 2):
        short = model_id.split("/")[-1]
        ws3.cell(row=row_idx, column=1, value=short).border = thin_border

        base_r = method_results["Baseline (FP16)"].get(model_id, {})
        base_lat = base_r.get("latency") if base_r.get("success") else None

        for col_idx, method_name in enumerate(method_names, 2):
            r = method_results[method_name].get(model_id, {})
            if r.get("success") and r.get("latency") is not None:
                lat = r["latency"]
                cell = ws3.cell(row=row_idx, column=col_idx, value=lat)
                cell.number_format = '0.000'
                if base_lat:
                    if lat <= base_lat:
                        cell.fill = green_fill
                    elif lat <= base_lat * 1.5:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill
            else:
                cell = ws3.cell(row=row_idx, column=col_idx, value="—")
                cell.fill = gray_fill
                cell.font = Font(color="808080")
            cell.border = thin_border

    ws3.column_dimensions["A"].width = 30
    for ci in range(2, len(method_names) + 2):
        ws3.column_dimensions[get_column_letter(ci)].width = 16
    ws3.freeze_panes = "B2"

    # ═══════════════════════════════════════════════════════════
    # Sheet 4: Per-Model × Method Grid (Peak Memory MB)
    # ═══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Peak Memory Grid")

    cols4 = ["Model"] + method_names
    style_header(ws4, cols4)

    for row_idx, model_id in enumerate(ALL_MODELS, 2):
        short = model_id.split("/")[-1]
        ws4.cell(row=row_idx, column=1, value=short).border = thin_border

        base_r = method_results["Baseline (FP16)"].get(model_id, {})
        base_mem = base_r.get("peak_memory") if base_r.get("success") else None

        for col_idx, method_name in enumerate(method_names, 2):
            r = method_results[method_name].get(model_id, {})
            if r.get("success") and r.get("peak_memory") is not None:
                mem = r["peak_memory"]
                cell = ws4.cell(row=row_idx, column=col_idx, value=mem)
                cell.number_format = '0.0'
                if base_mem:
                    if mem <= base_mem:
                        cell.fill = green_fill
                    elif mem <= base_mem * 1.2:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill
            else:
                cell = ws4.cell(row=row_idx, column=col_idx, value="—")
                cell.fill = gray_fill
                cell.font = Font(color="808080")
            cell.border = thin_border

    ws4.column_dimensions["A"].width = 30
    for ci in range(2, len(method_names) + 2):
        ws4.column_dimensions[get_column_letter(ci)].width = 16
    ws4.freeze_panes = "B2"

    # ═══════════════════════════════════════════════════════════
    # Sheet 5: Tier Analysis
    # ═══════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Tier Analysis")

    tier_data = {
        "100% Success": [],
        "High (>=70%)": [],
        "Moderate (45-69%)": [],
        "Low (<45%)": [],
    }

    for method_name, category, subdir, filter_fn, config in METHOD_DEFS:
        results = method_results[method_name]
        n_success = sum(1 for v in results.values() if v["success"])
        rate = (n_success / 19) * 100
        if rate == 100:
            tier_data["100% Success"].append((method_name, category, n_success, rate))
        elif rate >= 70:
            tier_data["High (>=70%)"].append((method_name, category, n_success, rate))
        elif rate >= 45:
            tier_data["Moderate (45-69%)"].append((method_name, category, n_success, rate))
        else:
            tier_data["Low (<45%)"].append((method_name, category, n_success, rate))

    tier_colors = {
        "100% Success": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
        "High (>=70%)": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
        "Moderate (45-69%)": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        "Low (<45%)": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    }
    tier_fonts = {
        "100% Success": Font(bold=True, color="FFFFFF", size=12),
        "High (>=70%)": Font(bold=True, color="000000", size=12),
        "Moderate (45-69%)": Font(bold=True, color="000000", size=12),
        "Low (<45%)": Font(bold=True, color="FFFFFF", size=12),
    }

    row = 1
    for tier_name, methods in tier_data.items():
        # Tier header
        cell = ws5.cell(row=row, column=1, value=f"{tier_name} ({len(methods)} methods)")
        cell.font = tier_fonts[tier_name]
        cell.fill = tier_colors[tier_name]
        cell.border = thin_border
        for ci in range(2, 5):
            c = ws5.cell(row=row, column=ci)
            c.fill = tier_colors[tier_name]
            c.border = thin_border
        row += 1

        # Sub-header
        for ci, name in enumerate(["Method", "Category", "Succeeded", "Rate (%)"], 1):
            c = ws5.cell(row=row, column=ci, value=name)
            c.font = Font(bold=True)
            c.border = thin_border
        row += 1

        for method_name, category, n_success, rate in methods:
            ws5.cell(row=row, column=1, value=method_name).border = thin_border
            ws5.cell(row=row, column=2, value=category).border = thin_border
            ws5.cell(row=row, column=3, value=f"{n_success}/19").border = thin_border
            c = ws5.cell(row=row, column=4, value=round(rate, 1))
            c.border = thin_border
            c.number_format = '0.0'
            row += 1

        row += 1  # blank row between tiers

    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 20
    ws5.column_dimensions["C"].width = 12
    ws5.column_dimensions["D"].width = 12

    # ═══════════════════════════════════════════════════════════
    # Sheet 6: Model Compatibility (which models work with most methods)
    # ═══════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Model Compatibility")

    cols6 = ["Model", "Methods Succeeded", "Out of 16", "Compatibility (%)", "Methods Failed"]
    style_header(ws6, cols6)

    method_names_no_baseline = [m[0] for m in METHOD_DEFS if m[0] != "Baseline (FP16)"]

    for row_idx, model_id in enumerate(ALL_MODELS, 2):
        short = model_id.split("/")[-1]
        n_success = 0
        failed_methods = []

        for method_name in method_names_no_baseline:
            r = method_results[method_name].get(model_id, {})
            if r.get("success"):
                n_success += 1
            else:
                failed_methods.append(method_name)

        n_total = len(method_names_no_baseline)  # 15 (excluding baseline)
        rate = (n_success / n_total) * 100

        ws6.cell(row=row_idx, column=1, value=short).border = thin_border
        c = ws6.cell(row=row_idx, column=2, value=n_success)
        c.border = thin_border
        c.font = bold_font
        ws6.cell(row=row_idx, column=3, value=n_total).border = thin_border

        c_rate = ws6.cell(row=row_idx, column=4, value=round(rate, 1))
        c_rate.border = thin_border
        c_rate.number_format = '0.0'
        if rate == 100:
            c_rate.fill = green_fill
        elif rate >= 70:
            c_rate.fill = PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid")
        elif rate >= 50:
            c_rate.fill = yellow_fill
        else:
            c_rate.fill = red_fill

        failed_str = ", ".join(failed_methods) if failed_methods else "None"
        ws6.cell(row=row_idx, column=5, value=failed_str).border = thin_border

    ws6.column_dimensions["A"].width = 32
    ws6.column_dimensions["B"].width = 18
    ws6.column_dimensions["C"].width = 10
    ws6.column_dimensions["D"].width = 18
    ws6.column_dimensions["E"].width = 80
    ws6.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Excel saved to: {output_path}")


if __name__ == "__main__":
    method_results = load_results()

    # Print quick summary
    total_s = 0
    total_t = 0
    for m_name, cat, subdir, fn, cfg in METHOD_DEFS:
        n = sum(1 for v in method_results[m_name].values() if v["success"])
        total_s += n
        total_t += 19
        print(f"  {m_name}: {n}/19")
    print(f"\n  TOTAL: {total_s}/{total_t} ({total_s/total_t*100:.1f}%)")

    output_path = os.path.join(os.path.dirname(__file__), "..", "VLM_Method_Summary.xlsx")
    create_excel(method_results, output_path)
