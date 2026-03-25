#!/usr/bin/env python3
"""Generate Excel file with only SUCCESSFUL CAD server results (non-zero accuracy, not all_failed)."""

import json
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS_DIR = os.path.join(os.path.dirname(__file__), "..", "results")

FAMILY_ORDER = [
    "ovis2", "internvl2.5", "qwen2.5-vl", "gemma3", "smolvlm", "lfm2-vl", "moondream2"
]

# Only CAD server method directories (exclude onnx which is empty)
CAD_DIRS = ["baseline", "ptq", "pruning", "wanda", "awq_gptq", "sparsegpt",
            "awp", "pact", "svd_llm", "palu", "casp_slim"]

def determine_method(data, subdir):
    if subdir == "baseline":
        return "Baseline (FP16)"
    elif subdir == "ptq":
        return "PTQ (BitsAndBytes)"
    elif subdir == "pruning":
        return "Magnitude Pruning"
    elif subdir == "wanda":
        return "Wanda"
    elif subdir == "sparsegpt":
        return "SparseGPT"
    elif subdir == "awq_gptq":
        m = data.get("method", "")
        return "AWQ (Sim. INT4)" if m == "awq" else "GPTQ (Sim. INT4)"
    elif subdir == "awp":
        return "AWP (Prune+Quant)"
    elif subdir == "pact":
        return "PACT (Prune+Merge)"
    elif subdir == "palu":
        return "PALU (KV Compression)"
    elif subdir == "svd_llm":
        return "SVD-LLM"
    elif subdir == "casp_slim":
        m = data.get("method", "")
        return "CASP (Mixed Quant)" if m == "casp" else "SLIM (SVD+Prune+Quant)"
    return subdir

def determine_config(data, subdir):
    if subdir == "baseline":
        return "FP16 (no compression)"
    elif subdir == "ptq":
        return f"{data.get('quant', '')} ({data.get('backend', '')})"
    elif subdir == "pruning":
        sp = data.get("target_sparsity", 0)
        return f"L1 Unstructured {int(sp*100)}%"
    elif subdir == "wanda":
        sp = data.get("target_sparsity", 0)
        return f"Wanda {int(sp*100)}%"
    elif subdir == "sparsegpt":
        sp = data.get("target_sparsity", 0)
        return f"SparseGPT {int(sp*100)}%"
    elif subdir == "awq_gptq":
        m = data.get("method", "").upper()
        return f"{m} Sim. INT4"
    elif subdir == "awp":
        sp = data.get("target_sparsity", 0)
        return f"Wanda {int(sp*100)}% + Sim. INT4"
    elif subdir == "pact":
        pr = data.get("prune_ratio", 0)
        mr = data.get("merge_ratio", 0)
        return f"Prune {int(pr*100)}% + Merge {int(mr*100)}%"
    elif subdir == "palu":
        e = data.get("energy_target", 0)
        return f"KV SVD energy={e}"
    elif subdir == "svd_llm":
        e = data.get("energy_target", 0)
        return f"MLP SVD energy={e}"
    elif subdir == "casp_slim":
        m = data.get("method", "")
        if m == "casp":
            return "Mixed 8/4-bit + LowRank QK"
        else:
            sp = data.get("target_sparsity", 0)
            rr = data.get("rank_ratio", 0)
            return f"SVD r={rr} + Prune {int(sp*100)}% + Quant"
    return ""

def family_sort_key(family):
    try:
        return FAMILY_ORDER.index(family)
    except ValueError:
        return len(FAMILY_ORDER)

def load_successful_results():
    records = []
    skipped = []

    for subdir in CAD_DIRS:
        dir_path = os.path.join(RESULTS_DIR, subdir)
        for json_path in sorted(glob.glob(os.path.join(dir_path, "*.json"))):
            if "gptq_comparison" in json_path:
                continue
            try:
                with open(json_path) as f:
                    data = json.load(f)
            except (json.JSONDecodeError, IOError):
                continue

            benchmarks = data.get("benchmarks", {})
            bench = benchmarks.get("vqav2", {})

            # Filter: skip if all_failed or zero/missing accuracy
            if bench.get("all_failed", False):
                skipped.append(os.path.basename(json_path))
                continue
            acc = bench.get("accuracy")
            if acc is None or acc == 0:
                skipped.append(os.path.basename(json_path))
                continue

            metrics = bench.get("metrics", {})
            method = determine_method(data, subdir)
            config = determine_config(data, subdir)

            record = {
                "Model": data.get("model_id", ""),
                "Family": data.get("family", ""),
                "Method": method,
                "Configuration": config,
                "Params (M)": data.get("num_params_M") or data.get("num_params_before_M", ""),
                "Params After (M)": data.get("num_params_after_M", ""),
                "GPU Mem Load (MB)": data.get("gpu_mem_load_mb", ""),
                "Compression Ratio": data.get("compression_ratio", ""),
                "Accuracy": bench.get("accuracy", ""),
                "Exact Match": metrics.get("exact_match", ""),
                "Contains": metrics.get("contains", ""),
                "Token F1": metrics.get("token_f1", ""),
                "BLEU": metrics.get("bleu", ""),
                "ROUGE-L": metrics.get("rouge_l", ""),
                "Avg Latency (s)": bench.get("avg_latency_s", ""),
                "Peak Memory (MB)": bench.get("peak_memory_mb", ""),
                "Avg Memory (MB)": bench.get("avg_memory_mb", ""),
                "Throughput (samples/s)": bench.get("throughput_sps", ""),
                "Avg Power (W)": bench.get("avg_power_w", ""),
                "GPU Util (%)": bench.get("avg_gpu_util_pct", ""),
                "Samples": bench.get("n_samples", ""),
                "Evaluated": bench.get("n_evaluated", ""),
                "Sparsity": data.get("target_sparsity") or data.get("actual_sparsity", ""),
                "Quant Bits": data.get("bits", ""),
            }
            records.append(record)

    return records, skipped

METHOD_ORDER = [
    "Baseline (FP16)", "PTQ (BitsAndBytes)", "Magnitude Pruning", "Wanda",
    "SparseGPT", "AWQ (Sim. INT4)", "GPTQ (Sim. INT4)", "AWP (Prune+Quant)",
    "PACT (Prune+Merge)", "PALU (KV Compression)", "SVD-LLM",
    "CASP (Mixed Quant)", "SLIM (SVD+Prune+Quant)"
]

def sort_key(r):
    fam = family_sort_key(r["Family"])
    try:
        meth = METHOD_ORDER.index(r["Method"])
    except ValueError:
        meth = len(METHOD_ORDER)
    return (fam, r["Model"], meth, r["Configuration"])

def apply_header_style(ws, columns):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

def create_excel(records, output_path):
    wb = Workbook()
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    accuracy_cols = {"Accuracy", "Exact Match", "Contains", "Token F1", "BLEU", "ROUGE-L", "Sparsity"}
    ratio_cols = {"Compression Ratio"}

    records.sort(key=sort_key)

    # ── Sheet 1: All Successful Results ──
    ws_all = wb.active
    ws_all.title = "All Successful Results"

    columns = [
        "Model", "Family", "Method", "Configuration",
        "Params (M)", "Params After (M)", "GPU Mem Load (MB)", "Compression Ratio",
        "Accuracy", "Exact Match", "Contains", "Token F1", "BLEU", "ROUGE-L",
        "Avg Latency (s)", "Peak Memory (MB)", "Avg Memory (MB)",
        "Throughput (samples/s)", "Avg Power (W)", "GPU Util (%)",
        "Samples", "Evaluated", "Sparsity", "Quant Bits"
    ]
    apply_header_style(ws_all, columns)

    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = record.get(col_name, "")
            cell = ws_all.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            if isinstance(val, float):
                if col_name in accuracy_cols:
                    cell.number_format = '0.0000'
                elif col_name in ratio_cols:
                    cell.number_format = '0.00'
                else:
                    cell.number_format = '0.0'

    for col_idx in range(1, len(columns) + 1):
        max_len = len(str(ws_all.cell(row=1, column=col_idx).value or ""))
        for row_idx in range(2, min(len(records) + 2, 30)):
            val = ws_all.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws_all.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 32)

    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(records)+1}"
    ws_all.freeze_panes = "A2"

    # ── Sheet 2: Accuracy Comparison (pivot) ──
    ws_acc = wb.create_sheet("Accuracy Comparison")

    # Build pivot: model -> method_config -> accuracy
    pivot = {}
    method_configs_seen = []
    model_family = {}
    baseline_acc = {}

    for r in records:
        model = r["Model"]
        method = r["Method"]
        config = r["Configuration"]
        key = f"{method} | {config}"
        acc = r["Accuracy"]
        if model not in pivot:
            pivot[model] = {}
        pivot[model][key] = acc
        if key not in method_configs_seen:
            method_configs_seen.append(key)
        model_family[model] = r["Family"]
        if method == "Baseline (FP16)":
            baseline_acc[model] = acc

    sum_cols = ["Model", "Family"] + method_configs_seen
    apply_header_style(ws_acc, sum_cols)

    sorted_models = sorted(pivot.keys(), key=lambda m: (family_sort_key(model_family.get(m, "")), m))

    for row_idx, model in enumerate(sorted_models, 2):
        ws_acc.cell(row=row_idx, column=1, value=model).border = thin_border
        ws_acc.cell(row=row_idx, column=2, value=model_family.get(model, "")).border = thin_border

        base = baseline_acc.get(model)

        for col_idx, mk in enumerate(method_configs_seen, 3):
            acc = pivot[model].get(mk, None)
            cell = ws_acc.cell(row=row_idx, column=col_idx, value=acc)
            cell.border = thin_border
            if isinstance(acc, (int, float)):
                cell.number_format = '0.0000'
                if base is not None and isinstance(base, (int, float)):
                    if acc >= base:
                        cell.fill = green_fill
                    elif acc >= base * 0.9:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill

    for col_idx in range(1, len(sum_cols) + 1):
        ws_acc.column_dimensions[get_column_letter(col_idx)].width = 24
    ws_acc.column_dimensions["A"].width = 40
    ws_acc.freeze_panes = "C2"

    # ── Sheet 3: Success Rate per Method ──
    ws_rate = wb.create_sheet("Method Success Rate")

    # 16 method variants × 19 models = 304 expected
    ALL_MODELS = [
        "HuggingFaceTB/SmolVLM-256M-Instruct", "LiquidAI/LFM2-VL-450M",
        "HuggingFaceTB/SmolVLM-500M-Instruct", "OpenGVLab/InternVL2_5-1B",
        "AIDC-AI/Ovis2-1B", "LiquidAI/LFM2-VL-1.6B", "OpenGVLab/InternVL2_5-2B",
        "AIDC-AI/Ovis2-2B", "vikhyatk/moondream2", "HuggingFaceTB/SmolVLM-Instruct",
        "LiquidAI/LFM2-VL-3B", "Qwen/Qwen2.5-VL-3B-Instruct",
        "OpenGVLab/InternVL2_5-4B", "AIDC-AI/Ovis2-4B", "google/gemma-3-4b-it",
        "Qwen/Qwen2.5-VL-7B-Instruct", "OpenGVLab/InternVL2_5-8B",
        "AIDC-AI/Ovis2-8B", "google/gemma-3-12b-it"
    ]

    # Count successes per method
    method_success = {}
    for r in records:
        m = r["Method"]
        c = r["Configuration"]
        key = f"{m} | {c}"
        if key not in method_success:
            method_success[key] = set()
        method_success[key].add(r["Model"])

    # Expected counts per method variant
    rate_cols = ["Method", "Configuration", "Successful", "Total Expected", "Success Rate (%)", "Failed Models"]
    apply_header_style(ws_rate, rate_cols)

    # Build the list from the CAD script: 16 variants × 19 models
    method_expected = {
        "Baseline (FP16) | FP16 (no compression)": 19,
        "PTQ (BitsAndBytes) | int8 (bnb)": 19,
        "PTQ (BitsAndBytes) | int4 (bnb)": 19,
        "Magnitude Pruning | L1 Unstructured 20%": 19,
        "Magnitude Pruning | L1 Unstructured 40%": 19,
        "Wanda | Wanda 20%": 19,
        "Wanda | Wanda 40%": 19,
        "AWQ (Sim. INT4) | AWQ Sim. INT4": 19,
        "GPTQ (Sim. INT4) | GPTQ Sim. INT4": 19,
        "SparseGPT | SparseGPT 50%": 19,
        "AWP (Prune+Quant) | Wanda 50% + Sim. INT4": 19,
        "PACT (Prune+Merge) | Prune 30% + Merge 20%": 19,
        "SVD-LLM | MLP SVD energy=0.95": 19,
        "PALU (KV Compression) | KV SVD energy=0.95": 19,
        "CASP (Mixed Quant) | Mixed 8/4-bit + LowRank QK": 19,
        "SLIM (SVD+Prune+Quant) | SVD r=0.3 + Prune 50% + Quant": 19,
    }

    row_idx = 2
    for key, expected in method_expected.items():
        parts = key.split(" | ", 1)
        method_name = parts[0]
        config_name = parts[1] if len(parts) > 1 else ""
        successful_models = method_success.get(key, set())
        n_success = len(successful_models)
        rate = (n_success / expected) * 100

        failed_models = sorted(set(ALL_MODELS) - successful_models)
        failed_str = ", ".join([m.split("/")[-1] for m in failed_models]) if failed_models else "None"

        ws_rate.cell(row=row_idx, column=1, value=method_name).border = thin_border
        ws_rate.cell(row=row_idx, column=2, value=config_name).border = thin_border
        ws_rate.cell(row=row_idx, column=3, value=n_success).border = thin_border
        ws_rate.cell(row=row_idx, column=4, value=expected).border = thin_border
        cell_rate = ws_rate.cell(row=row_idx, column=5, value=rate)
        cell_rate.border = thin_border
        cell_rate.number_format = '0.0'
        if rate == 100:
            cell_rate.fill = green_fill
        elif rate >= 50:
            cell_rate.fill = yellow_fill
        else:
            cell_rate.fill = red_fill
        ws_rate.cell(row=row_idx, column=6, value=failed_str).border = thin_border
        row_idx += 1

    # Totals row
    total_success = sum(len(v) for v in method_success.values())
    total_expected = sum(method_expected.values())
    ws_rate.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True)
    ws_rate.cell(row=row_idx, column=3, value=total_success).font = Font(bold=True)
    ws_rate.cell(row=row_idx, column=4, value=total_expected).font = Font(bold=True)
    ws_rate.cell(row=row_idx, column=5, value=(total_success/total_expected)*100).font = Font(bold=True)
    ws_rate.cell(row=row_idx, column=5).number_format = '0.0'

    ws_rate.column_dimensions["A"].width = 30
    ws_rate.column_dimensions["B"].width = 35
    ws_rate.column_dimensions["C"].width = 12
    ws_rate.column_dimensions["D"].width = 16
    ws_rate.column_dimensions["E"].width = 16
    ws_rate.column_dimensions["F"].width = 80

    # ── Sheet 4: Methods Overview ──
    ws_methods = wb.create_sheet("Methods Overview")
    method_info = [
        ["Method", "Category", "Description", "Key Parameters"],
        ["Baseline (FP16)", "None", "Original model, no compression", "FP16 weights"],
        ["PTQ (BitsAndBytes)", "Quantization", "Post-training quantization via BitsAndBytes", "INT4 / INT8"],
        ["AWQ (Sim. INT4)", "Quantization", "Activation-aware weight quantization (simulated)", "4-bit, per-channel groups"],
        ["GPTQ (Sim. INT4)", "Quantization", "GPTQ quantization (simulated)", "4-bit, per-channel groups"],
        ["Magnitude Pruning", "Pruning", "L1 unstructured magnitude pruning", "Sparsity: 20%, 40%"],
        ["Wanda", "Pruning", "Pruning by weights and activations", "Sparsity: 20%, 40%"],
        ["SparseGPT", "Pruning", "One-shot pruning with Hessian-based updates", "Sparsity: 50%"],
        ["AWP (Prune+Quant)", "Combined", "Wanda pruning + simulated INT4 quantization", "Sparsity 50% + INT4"],
        ["PACT (Prune+Merge)", "Token Compression", "Progressive token pruning and merging", "Prune 30% + Merge 20%"],
        ["PALU (KV Compression)", "Low-Rank", "SVD-based KV cache compression", "Energy target 0.95"],
        ["SVD-LLM", "Low-Rank", "SVD decomposition of MLP layers", "Energy target 0.95"],
        ["CASP (Mixed Quant)", "Mixed Precision", "Channel-aware sensitivity-based mixed precision", "8/4-bit + LowRank QK"],
        ["SLIM (SVD+Prune+Quant)", "Combined", "SVD + pruning + quantization pipeline", "Rank 0.3 + Sparsity 50% + Quant"],
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_i, row_data in enumerate(method_info, 1):
        for col_i, val in enumerate(row_data, 1):
            cell = ws_methods.cell(row=row_i, column=col_i, value=val)
            cell.border = thin_border
            if row_i == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

    ws_methods.column_dimensions["A"].width = 30
    ws_methods.column_dimensions["B"].width = 18
    ws_methods.column_dimensions["C"].width = 55
    ws_methods.column_dimensions["D"].width = 35

    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")
    print(f"Total successful records: {len(records)}")

if __name__ == "__main__":
    records, skipped = load_successful_results()
    print(f"Successful: {len(records)}, Skipped (failed/zero-acc): {len(skipped)}")
    output_path = os.path.join(os.path.dirname(__file__), "..", "VLM_CAD_Successful_Results.xlsx")
    create_excel(records, output_path)
