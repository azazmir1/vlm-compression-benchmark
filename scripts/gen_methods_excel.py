#!/usr/bin/env python3
"""Generate VLM Compression Methods Reference Excel with GitHub links, paper links, etc."""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read source data
src = openpyxl.load_workbook("VLM_Memory_Optimization_Papers.xlsx")
ws_src = src.active
rows_data = []
for row in ws_src.iter_rows(min_row=2, values_only=True):
    if row[0]:
        rows_data.append(list(row))

# GitHub repos mapping
github_repos = {
    "AWQ": "https://github.com/mit-han-lab/llm-awq",
    "GPTQ": "https://github.com/IST-DASLab/gptq",
    "BnB INT8 / INT4 (NF4)": "https://github.com/bitsandbytes-foundation/bitsandbytes",
    "Q-VLM": "https://github.com/ChangyuanWang17/QVLM",
    "LeanQuant": "https://github.com/LeanModels/LeanQuant",
    "SpinQuant": "https://github.com/facebookresearch/SpinQuant",
    "Wanda": "https://github.com/locuslab/wanda",
    "SparseGPT": "https://github.com/IST-DASLab/sparsegpt",
    "Magnitude Pruning": "N/A (classic method)",
    "M-Wanda": "N/A (no public repo)",
    "AWP": "No official repo (MERL, ICML 2025)",
    "PACT": "https://github.com/orailix/PACT",
    "ATP-LLaVA": "No public repo yet",
    "LLaVA-Mini": "https://github.com/ictnlp/LLaVA-Mini",
    "Dynamic-LLaVA": "https://github.com/Osilly/dynamic_llava",
    "DivPrune": "https://github.com/vbdi/divprune",
    "DyCoke": "https://github.com/KD-TAO/DyCoke",
    "LLaVA-KD": "https://github.com/Fantasyele/LLaVA-KD",
    "LLaVA-MoD": "https://github.com/shufangxun/LLaVA-MoD",
    "TinyLLaVA": "https://github.com/TinyLLaVA/TinyLLaVA_Factory",
    "SVD-LLM": "https://github.com/AIoT-MLSys-Lab/SVD-LLM",
    "CALDERA": "https://github.com/pilancilab/caldera",
    "PALU": "https://github.com/shadowpa0327/Palu",
    "SLIM": "https://github.com/Paramathic/slim",
    "CASP": "https://github.com/vbdi/casp",
    "SmolVLM": "https://github.com/huggingface/smollm",
    "MiniCPM-V": "https://github.com/OpenBMB/MiniCPM-V",
    "N/A": "",
}

tool_repos = {
    "AWQ": "AutoAWQ: https://github.com/casper-hansen/AutoAWQ",
    "GPTQ": "AutoGPTQ: https://github.com/AutoGPTQ/AutoGPTQ",
}

our_status = {
    "AWQ": "Code done (compression/ptq/run_awq.py) - ExllamaV2 fails on Jetson",
    "GPTQ": "Code done (compression/ptq/run_gptq.py) - Works on Jetson via Triton",
    "BnB INT8 / INT4 (NF4)": "DONE on A6000 - BLOCKED on Jetson (CUDA kernel crash)",
    "Q-VLM": "Not implemented",
    "LeanQuant": "Not implemented",
    "SpinQuant": "Not implemented",
    "Wanda": "DONE - 77 results on A6000 (compression/pruning/run_wanda.py)",
    "SparseGPT": "Code done (compression/pruning/run_sparsegpt.py) - cuSolver crash on Jetson",
    "Magnitude Pruning": "DONE - All models at 20%/40% (compression/pruning/run_pruning.py)",
    "M-Wanda": "Skipped (English-only models)",
    "AWP": "Code done (compression/combined/run_awp.py)",
    "PACT": "Code done, faithful to CVPR 2025 paper (compression/pact/run_pact.py)",
    "ATP-LLaVA": "Not implemented (requires training)",
    "LLaVA-Mini": "Skipped (LLaVA-specific architecture)",
    "Dynamic-LLaVA": "Not implemented",
    "DivPrune": "Not implemented",
    "DyCoke": "Skipped (video-specific)",
    "LLaVA-KD": "Skipped (requires training from scratch)",
    "LLaVA-MoD": "Skipped (LLaVA-specific + training)",
    "TinyLLaVA": "Skipped (architecture design, not compression)",
    "SVD-LLM": "Code done (compression/lowrank/run_svd_llm.py)",
    "CALDERA": "Not implemented",
    "PALU": "Code done (compression/palu/run_palu.py)",
    "SLIM": "Code done (compression/casp_slim/run_casp_slim.py)",
    "CASP": "Code done (compression/casp_slim/run_casp_slim.py)",
    "SmolVLM": "In benchmark (baseline models)",
    "MiniCPM-V": "Not in benchmark",
    "N/A": "Reference only",
    "LLMC+ (benchmark)": "Reference only",
}

jetson_compat = {
    "AWQ": "Partial - ExllamaV2 kernel divisibility issues",
    "GPTQ": "YES - Triton backend works",
    "BnB INT8 / INT4 (NF4)": "NO - CUDA kernel crash on aarch64",
    "Q-VLM": "Unknown",
    "LeanQuant": "Unknown",
    "SpinQuant": "Unknown",
    "Wanda": "YES - pure PyTorch",
    "SparseGPT": "NO - cuSolver symbol missing",
    "Magnitude Pruning": "YES - pure PyTorch",
    "M-Wanda": "N/A",
    "AWP": "Partial - Wanda works, BnB quant blocked",
    "PACT": "YES - pure PyTorch",
    "ATP-LLaVA": "Unknown",
    "LLaVA-Mini": "N/A",
    "Dynamic-LLaVA": "Unknown",
    "DivPrune": "Likely YES (training-free)",
    "DyCoke": "N/A",
    "LLaVA-KD": "N/A",
    "LLaVA-MoD": "N/A",
    "TinyLLaVA": "N/A",
    "SVD-LLM": "YES - pure PyTorch",
    "CALDERA": "Unknown",
    "PALU": "YES - pure PyTorch",
    "SLIM": "Partial - depends on sub-methods",
    "CASP": "Partial - depends on sub-methods",
    "SmolVLM": "YES (models load on Jetson)",
    "MiniCPM-V": "Unknown",
    "N/A": "N/A",
    "LLMC+ (benchmark)": "N/A",
}

# Styles
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
cat_fills = {
    "Weight Quantization": PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid"),
    "Weight Pruning": PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid"),
    "Weight Pruning + Quantization": PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"),
    "Visual Token Compression": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Knowledge Distillation": PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid"),
    "Knowledge Distillation + MoE": PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid"),
    "Efficient Architecture": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    "Low-Rank Decomposition": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "Low-Rank + Quantization": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "KV-Cache Compression": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "Combined (Quant+Prune+LowRank)": PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"),
    "Combined (VLM-specific)": PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"),
    "Survey": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    "Survey / Benchmark": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}
yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
link_font = Font(color="0563C1", underline="single")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_hdr(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border


def auto_w(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 3, 55)


wb = Workbook()

# ===== Sheet 1: Methods Overview =====
ws1 = wb.active
ws1.title = "Methods Overview"
hdrs = [
    "#", "Method", "Category", "Paper Title", "Authors", "Venue", "Year",
    "Key Idea", "Memory Saving", "GitHub Repo", "Tool Library",
    "Paper Link", "Our Implementation Status", "Jetson Compatible", "Recommendation",
]
for c, h in enumerate(hdrs, 1):
    ws1.cell(row=1, column=c, value=h)
style_hdr(ws1, len(hdrs))

for i, rd in enumerate(rows_data):
    row = i + 2
    paper, authors, category, method, venue, year, target, key_idea, mem_saving, status, recommendation, link = rd[:12]

    ws1.cell(row=row, column=1, value=i + 1)
    ws1.cell(row=row, column=2, value=method or "")
    ws1.cell(row=row, column=3, value=category or "")
    ws1.cell(row=row, column=4, value=paper or "")
    ws1.cell(row=row, column=5, value=authors or "")
    ws1.cell(row=row, column=6, value=venue or "")
    ws1.cell(row=row, column=7, value=year)
    ws1.cell(row=row, column=8, value=key_idea or "")
    ws1.cell(row=row, column=9, value=mem_saving or "")

    gh = github_repos.get(method, "")
    ws1.cell(row=row, column=10, value=gh)
    if gh.startswith("https://"):
        ws1.cell(row=row, column=10).font = link_font

    tl = tool_repos.get(method, "")
    ws1.cell(row=row, column=11, value=tl)

    ws1.cell(row=row, column=12, value=link or "")
    if link and str(link).startswith("http"):
        ws1.cell(row=row, column=12).font = link_font

    ws1.cell(row=row, column=13, value=our_status.get(method, ""))

    jc = jetson_compat.get(method, "")
    ws1.cell(row=row, column=14, value=jc)
    jc_cell = ws1.cell(row=row, column=14)
    if jc.startswith("YES"):
        jc_cell.fill = yes_fill
    elif jc.startswith("NO"):
        jc_cell.fill = no_fill
    elif jc.startswith("Partial"):
        jc_cell.fill = partial_fill

    ws1.cell(row=row, column=15, value=recommendation or "")

    cfill = cat_fills.get(category, None)
    if cfill:
        ws1.cell(row=row, column=3).fill = cfill

    for c in range(1, len(hdrs) + 1):
        ws1.cell(row=row, column=c).border = thin_border
        ws1.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")

auto_w(ws1)
ws1.column_dimensions["D"].width = 50
ws1.column_dimensions["H"].width = 55
ws1.column_dimensions["J"].width = 45
ws1.column_dimensions["M"].width = 50
ws1.column_dimensions["O"].width = 55

# ===== Sheet 2: Quick Reference =====
ws2 = wb.create_sheet("Quick Reference")
hdrs2 = ["Method", "Category", "Venue/Year", "GitHub Repo", "Paper Link", "Jetson OK?", "Our Status"]
for c, h in enumerate(hdrs2, 1):
    ws2.cell(row=1, column=c, value=h)
style_hdr(ws2, len(hdrs2))

methods_only = [rd for rd in rows_data if rd[2] not in ("Survey", "Survey / Benchmark")]
for i, rd in enumerate(methods_only):
    row = i + 2
    paper, authors, category, method, venue, year, target, key_idea, mem_saving, status, recommendation, link = rd[:12]

    ws2.cell(row=row, column=1, value=method or "")
    ws2.cell(row=row, column=2, value=category or "")
    vy = "{} {}".format(venue, year) if venue else str(year)
    ws2.cell(row=row, column=3, value=vy)

    gh = github_repos.get(method, "")
    ws2.cell(row=row, column=4, value=gh)
    if gh.startswith("https://"):
        ws2.cell(row=row, column=4).font = link_font

    ws2.cell(row=row, column=5, value=link or "")
    if link and str(link).startswith("http"):
        ws2.cell(row=row, column=5).font = link_font

    jc = jetson_compat.get(method, "")
    ws2.cell(row=row, column=6, value=jc)
    jc_cell = ws2.cell(row=row, column=6)
    if jc.startswith("YES"):
        jc_cell.fill = yes_fill
    elif jc.startswith("NO"):
        jc_cell.fill = no_fill
    elif jc.startswith("Partial"):
        jc_cell.fill = partial_fill

    ws2.cell(row=row, column=7, value=our_status.get(method, ""))

    cfill = cat_fills.get(category, None)
    if cfill:
        ws2.cell(row=row, column=2).fill = cfill

    for c in range(1, len(hdrs2) + 1):
        ws2.cell(row=row, column=c).border = thin_border

auto_w(ws2)

# ===== Sheet 3: By Category =====
ws3 = wb.create_sheet("By Category")
hdrs3 = ["Category", "Method", "Memory Saving", "GitHub", "Jetson OK?", "Our Status"]
for c, h in enumerate(hdrs3, 1):
    ws3.cell(row=1, column=c, value=h)
style_hdr(ws3, len(hdrs3))

by_cat = sorted(methods_only, key=lambda r: (r[2] or "", r[3] or ""))
for i, rd in enumerate(by_cat):
    row = i + 2
    paper, authors, category, method, venue, year, target, key_idea, mem_saving, status, recommendation, link = rd[:12]

    ws3.cell(row=row, column=1, value=category or "")
    ws3.cell(row=row, column=2, value=method or "")
    ws3.cell(row=row, column=3, value=mem_saving or "")

    gh = github_repos.get(method, "")
    ws3.cell(row=row, column=4, value=gh)
    if gh.startswith("https://"):
        ws3.cell(row=row, column=4).font = link_font

    jc = jetson_compat.get(method, "")
    ws3.cell(row=row, column=5, value=jc)
    jc_cell = ws3.cell(row=row, column=5)
    if jc.startswith("YES"):
        jc_cell.fill = yes_fill
    elif jc.startswith("NO"):
        jc_cell.fill = no_fill
    elif jc.startswith("Partial"):
        jc_cell.fill = partial_fill

    ws3.cell(row=row, column=6, value=our_status.get(method, ""))

    cfill = cat_fills.get(category, None)
    if cfill:
        ws3.cell(row=row, column=1).fill = cfill

    for c in range(1, len(hdrs3) + 1):
        ws3.cell(row=row, column=c).border = thin_border

auto_w(ws3)

# ===== Sheet 4: Surveys & References =====
ws4 = wb.create_sheet("Surveys & References")
surveys = [rd for rd in rows_data if rd[2] in ("Survey", "Survey / Benchmark", "Efficient Architecture")]
hdrs4 = ["Paper Title", "Authors", "Type", "Venue/Year", "Paper Link", "GitHub", "Notes"]
for c, h in enumerate(hdrs4, 1):
    ws4.cell(row=1, column=c, value=h)
style_hdr(ws4, len(hdrs4))

for i, rd in enumerate(surveys):
    row = i + 2
    paper, authors, category, method, venue, year, target, key_idea, mem_saving, status, recommendation, link = rd[:12]
    ws4.cell(row=row, column=1, value=paper or "")
    ws4.cell(row=row, column=2, value=authors or "")
    ws4.cell(row=row, column=3, value=category or "")
    ws4.cell(row=row, column=4, value="{} {}".format(venue, year))
    ws4.cell(row=row, column=5, value=link or "")
    if link and str(link).startswith("http"):
        ws4.cell(row=row, column=5).font = link_font
    gh = github_repos.get(method, "")
    ws4.cell(row=row, column=6, value=gh)
    if gh.startswith("https://"):
        ws4.cell(row=row, column=6).font = link_font
    ws4.cell(row=row, column=7, value=recommendation or "")
    for c in range(1, len(hdrs4) + 1):
        ws4.cell(row=row, column=c).border = thin_border

auto_w(ws4)

outpath = "results/VLM_Compression_Methods_Reference.xlsx"
wb.save(outpath)
print("Saved:", outpath)
print("  Sheet 1: Methods Overview -", len(rows_data), "papers")
print("  Sheet 2: Quick Reference -", len(methods_only), "methods")
print("  Sheet 3: By Category - grouped by type")
print("  Sheet 4: Surveys & References -", len(surveys), "papers")
