#!/usr/bin/env python3
"""Create Excel sheet: all model families, variants, params, Jetson FP16 status, and compression results."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "VLM Models - Jetson Status"

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
cat2_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
error_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
compressed_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Headers — 5 metrics per compression method
methods = ["HQQ INT4", "PyTorch INT8", "PyTorch INT4", "GPTQ INT4"]
metric_names = ["EM", "C", "F1", "BLEU", "R-L"]
base_headers = ["Family", "Model Variant", "Parameters (M)", "Jetson FP16 Status",
                "GPU Memory (MB)", "Category", "Notes"]
headers = list(base_headers)
for m in methods:
    for met in metric_names:
        headers.append(f"{m}\n{met}")

# Data: [family, model, params_M, jetson_status, mem_mb, category, notes,
#        hqq_em, hqq_c, hqq_f1, hqq_bleu, hqq_rl,
#        int8_em, int8_c, int8_f1, int8_bleu, int8_rl,
#        int4_em, int4_c, int4_f1, int4_bleu, int4_rl,
#        gptq_em, gptq_c, gptq_f1, gptq_bleu, gptq_rl]
_na = ["--"]*5  # no results for a method

models = [
    # SmolVLM
    ["SmolVLM", "SmolVLM-256M-Instruct", 260, "PASS", 333, "Below Ceiling", "Runs fine in FP16",
     *_na, *_na, *_na, *_na],
    ["SmolVLM", "SmolVLM-500M-Instruct", 508, "PASS", 537, "Ceiling", "Largest that runs usably",
     *_na, *_na, *_na, *_na],
    ["SmolVLM", "SmolVLM-2.2B-Instruct", 2200, "OOM_LOAD", "--", "Category 1", "Cannot load in FP16",
     0.647, 0.680, 0.707, 0.695, 0.707,
     0.640, 0.680, 0.693, 0.687, 0.693,
     0.580, 0.600, 0.613, 0.607, 0.613,
     0.473, 0.540, 0.577, 0.565, 0.577],

    # LFM2-VL
    ["LFM2-VL", "LFM2-VL-450M", 450, "PASS", 462, "Below Ceiling", "Runs fine in FP16",
     *_na, *_na, *_na, *_na],
    ["LFM2-VL", "LFM2-VL-1.6B", 1600, "MEM_CRITICAL", 2916, "Category 2", "Loads but system nearly OOM, unusable",
     0.820, 0.820, 0.820, 0.820, 0.820,
     0.813, 0.820, 0.835, 0.830, 0.835,
     0.820, 0.840, 0.855, 0.850, 0.855,
     0.820, 0.840, 0.855, 0.850, 0.855],
    ["LFM2-VL", "LFM2-VL-3B", 3000, "OOM_LOAD", "--", "Category 1", "Cannot load — exceeds memory",
     *_na, *_na, *_na, *_na],

    # InternVL2.5
    ["InternVL2.5", "InternVL2.5-1B", 1000, "PASS", 1410, "Ceiling", "Largest that runs usably",
     *_na, *_na, *_na, *_na],
    ["InternVL2.5", "InternVL2.5-2B", 2200, "OOM_LOAD", "--", "Category 1", "Cannot load — exceeds memory",
     *_na, *_na, *_na, *_na],
    ["InternVL2.5", "InternVL2.5-4B", 3700, "OOM_LOAD", "--", "Category 1", "Cannot load — exceeds memory",
     *_na, *_na, *_na, *_na],
    ["InternVL2.5", "InternVL2.5-8B", 8100, "OOM_LOAD", "--", "Category 1", "Cannot load — exceeds memory",
     *_na, *_na, *_na, *_na],

    # Qwen2.5-VL
    ["Qwen2.5-VL", "Qwen2.5-VL-3B-Instruct", 3800, "OOM_LOAD", "--", "Category 1", "Cannot load — exceeds memory",
     "OOM", "OOM", "OOM", "OOM", "OOM",
     0.800, 0.860, 0.873, 0.867, 0.873,
     0.827, 0.900, 0.913, 0.907, 0.913,
     0.807, 0.880, 0.893, 0.887, 0.893],
    ["Qwen2.5-VL", "Qwen2.5-VL-7B-Instruct", 8300, "OOM_LOAD", "--", "Category 1", "Cannot load — exceeds memory",
     *_na, *_na, *_na, *_na],

    # Ovis2
    ["Ovis2", "Ovis2-1B", 1000, "ERROR", 1892, "Error", "Remote-code config conflict, not memory",
     *_na, *_na, *_na, *_na],
    ["Ovis2", "Ovis2-2B", 2200, "OOM_LOAD", "--", "Category 1", "Cannot load — exceeds memory",
     *_na, *_na, *_na, *_na],
    ["Ovis2", "Ovis2-4B", 4300, "OOM_LOAD", "--", "Category 1", "Cannot load — exceeds memory",
     *_na, *_na, *_na, *_na],
    ["Ovis2", "Ovis2-8B", 8900, "OOM_LOAD", "--", "Category 1", "Cannot load — exceeds memory",
     *_na, *_na, *_na, *_na],

    # Moondream
    ["Moondream", "moondream2", 2000, "MEM_CRITICAL", 2822, "Category 2", "Loads but system nearly OOM, unusable",
     *_na, *_na, *_na, *_na],

    # Gemma 3
    ["Gemma 3", "gemma-3-4b-it", 4300, "OOM_LOAD", "--", "Category 1", "Cannot load — exceeds memory",
     *_na, *_na, *_na, *_na],
    ["Gemma 3", "gemma-3-12b-it", 12200, "OOM_LOAD", "--", "Category 1", "Cannot load — exceeds memory",
     *_na, *_na, *_na, *_na],

    # FastVLM
    ["FastVLM", "FastVLM-0.5B", 500, "PASS", 1989, "Ceiling", "Runs but high memory usage",
     *_na, *_na, *_na, *_na],
    ["FastVLM", "FastVLM-1.5B", 1500, "OOM_LOAD", "--", "Category 1", "Cannot load — exceeds memory",
     *_na, *_na, *_na, *_na],
]

# Write headers: row 1 = method group names (merged), row 2 = metric sub-headers
n_base = len(base_headers)
for col, h in enumerate(base_headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

for mi, m in enumerate(methods):
    start_col = n_base + 1 + mi * len(metric_names)
    end_col = start_col + len(metric_names) - 1
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    cell = ws.cell(row=1, column=start_col, value=m)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border
    for si, met in enumerate(metric_names):
        sub_cell = ws.cell(row=2, column=start_col + si, value=met)
        sub_cell.font = header_font
        sub_cell.fill = header_fill
        sub_cell.alignment = Alignment(horizontal='center', wrap_text=True)
        sub_cell.border = thin_border

# Write data (row 3 onwards)
for row_idx, data in enumerate(models, 3):
    n_base = len(base_headers)
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if col_idx >= 3 else 'left',
                                   wrap_text=True)

    # Color code status column
    status = data[3]
    status_cell = ws.cell(row=row_idx, column=4)
    cat_cell = ws.cell(row=row_idx, column=6)

    if status == "PASS":
        status_cell.fill = pass_fill
        cat_cell.fill = pass_fill
    elif status == "OOM_LOAD":
        status_cell.fill = fail_fill
        cat_cell.fill = fail_fill
    elif status == "MEM_CRITICAL":
        status_cell.fill = cat2_fill
        cat_cell.fill = cat2_fill
    elif status == "ERROR":
        status_cell.fill = error_fill
        cat_cell.fill = error_fill

    # Color code compression metric cells (columns 8+)
    for col_idx in range(n_base + 1, len(data) + 1):
        val = data[col_idx - 1]
        cell = ws.cell(row=row_idx, column=col_idx)
        if val != "--":
            if val == "OOM":
                cell.fill = fail_fill
            else:
                cell.fill = compressed_fill

# Column widths
base_widths = [14, 28, 16, 18, 16, 14, 40]
metric_width = 8
col_widths = base_widths + [metric_width] * (len(methods) * len(metric_names))
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze top row
ws.freeze_panes = 'A3'

# Add summary sheet
ws2 = wb.create_sheet("Summary")
ws2.cell(row=1, column=1, value="Category").font = Font(bold=True)
ws2.cell(row=1, column=2, value="Count").font = Font(bold=True)
ws2.cell(row=1, column=3, value="Description").font = Font(bold=True)

summary = [
    ["PASS (Runs on Jetson)", sum(1 for m in models if m[3] == "PASS"), "Loads and runs inference in FP16"],
    ["Category 1 (Cannot Load)", sum(1 for m in models if m[3] == "OOM_LOAD"), "Exceeds 8GB unified memory, OOM at load"],
    ["Category 2 (Loads, Unusable)", sum(1 for m in models if m[3] == "MEM_CRITICAL"), "Loads but crashes or too slow during inference"],
    ["Error (Non-memory)", sum(1 for m in models if m[3] == "ERROR"), "Fails due to code/config issues, not memory"],
    ["Total Models", len(models), "Across 9 VLM families"],
]
for r, row in enumerate(summary, 2):
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)

ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 50

out = "/home/cselab/vlm-compression-benchmark/results/_excel/VLM_Model_Families_Jetson_Status.xlsx"
wb.save(out)
print(f"Saved to {out}")

# Print summary
pass_count = sum(1 for m in models if m[3] == "PASS")
oom_count = sum(1 for m in models if m[3] == "OOM_LOAD")
cat2_count = sum(1 for m in models if m[3] == "MEM_CRITICAL")
err_count = sum(1 for m in models if m[3] == "ERROR")
print(f"\nSummary:")
print(f"  PASS (runs on Jetson FP16): {pass_count}")
print(f"  Category 1 (OOM, cannot load): {oom_count}")
print(f"  Category 2 (loads, unusable): {cat2_count}")
print(f"  Error (non-memory): {err_count}")
print(f"  Total: {len(models)}")
