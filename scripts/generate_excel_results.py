#!/usr/bin/env python3
"""Generate a comprehensive Excel file with all VLM compression benchmark results."""

import json
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS_DIR = os.path.join(os.path.dirname(__file__), "..", "results")

# Model family display order
FAMILY_ORDER = [
    "ovis2", "internvl2.5", "qwen2.5-vl", "gemma3", "smolvlm", "lfm2-vl", "moondream2"
]

def load_all_results():
    """Load all JSON result files and return a list of parsed records."""
    records = []

    for json_path in sorted(glob.glob(os.path.join(RESULTS_DIR, "**", "*.json"), recursive=True)):
        if "gptq_comparison" in json_path:
            continue
        try:
            with open(json_path) as f:
                data = json.load(f)
        except (json.JSONDecodeError, IOError):
            continue

        # Determine method category from directory
        subdir = os.path.basename(os.path.dirname(json_path))

        # Extract common fields
        model_id = data.get("model_id", "")
        family = data.get("family", "")
        method = determine_method(data, subdir)
        config = determine_config(data, subdir)

        # Extract benchmark metrics
        benchmarks = data.get("benchmarks", {})
        bench = benchmarks.get("vqav2", {})
        metrics = bench.get("metrics", {})

        record = {
            "Model": model_id,
            "Family": family,
            "Method": method,
            "Configuration": config,
            "Params (M)": data.get("num_params_M") or data.get("num_params_before_M", ""),
            "Params After (M)": data.get("num_params_after_M", ""),
            "GPU Mem Load (MB)": data.get("gpu_mem_load_mb", ""),
            "Compression Ratio": data.get("compression_ratio", ""),
            # VQAv2 benchmark
            "Accuracy": bench.get("accuracy", ""),
            "Exact Match": metrics.get("exact_match", ""),
            "Contains": metrics.get("contains", ""),
            "Token F1": metrics.get("token_f1", ""),
            "BLEU": metrics.get("bleu", ""),
            "ROUGE-L": metrics.get("rouge_l", ""),
            "Avg Latency (s)": bench.get("avg_latency_s", ""),
            "Peak Memory (MB)": bench.get("peak_memory_mb", ""),
            "Avg Memory (MB)": bench.get("avg_memory_mb", ""),
            "Throughput (samples/s)": bench.get("throughput_sps", ""),
            "Avg Power (W)": bench.get("avg_power_w", ""),
            "GPU Util (%)": bench.get("avg_gpu_util_pct", ""),
            "Samples": bench.get("n_samples", ""),
            "Evaluated": bench.get("n_evaluated", ""),
            "Skipped": bench.get("n_skipped", ""),
            # Extra method-specific
            "Sparsity": data.get("target_sparsity") or data.get("actual_sparsity", ""),
            "Quant Bits": data.get("bits", ""),
        }
        records.append(record)

    return records

def determine_method(data, subdir):
    """Determine the compression method name."""
    method_map = {
        "baseline": "Baseline (FP16)",
        "ptq": "PTQ (BitsAndBytes)",
        "pruning": "Magnitude Pruning",
        "wanda": "Wanda",
        "sparsegpt": "SparseGPT",
        "awq_gptq": data.get("method", "").upper() if data.get("method") else "AWQ/GPTQ",
        "awp": "AWP (Prune+Quant)",
        "pact": "PACT (Prune+Merge)",
        "palu": "PALU (KV Compression)",
        "svd_llm": "SVD-LLM",
        "casp_slim": data.get("method", "").upper() if data.get("method") else "CASP/SLIM",
        "onnx": "ONNX",
    }
    method = method_map.get(subdir, subdir)
    if subdir == "awq_gptq":
        m = data.get("method", "")
        method = f"AWQ (Sim. INT4)" if m == "awq" else f"GPTQ (Sim. INT4)"
    if subdir == "casp_slim":
        m = data.get("method", "")
        method = "CASP (Mixed Quant)" if m == "casp" else "SLIM (SVD+Prune+Quant)"
    return method

def determine_config(data, subdir):
    """Determine a human-readable configuration string."""
    if subdir == "baseline":
        return "FP16 (no compression)"
    elif subdir == "ptq":
        return f"{data.get('quant', '')} ({data.get('backend', '')})"
    elif subdir == "pruning":
        sp = data.get("target_sparsity", 0)
        return f"L1 Unstructured {int(sp*100)}%"
    elif subdir == "wanda":
        sp = data.get("target_sparsity", 0)
        return f"Wanda {int(sp*100)}%"
    elif subdir == "sparsegpt":
        sp = data.get("target_sparsity", 0)
        return f"SparseGPT {int(sp*100)}%"
    elif subdir == "awq_gptq":
        m = data.get("method", "").upper()
        return f"{m} Sim. INT4"
    elif subdir == "awp":
        sp = data.get("target_sparsity", 0)
        return f"Wanda {int(sp*100)}% + Sim. INT4"
    elif subdir == "pact":
        pr = data.get("prune_ratio", 0)
        mr = data.get("merge_ratio", 0)
        return f"Prune {int(pr*100)}% + Merge {int(mr*100)}%"
    elif subdir == "palu":
        e = data.get("energy_target", 0)
        return f"KV SVD energy={e}"
    elif subdir == "svd_llm":
        e = data.get("energy_target", 0)
        return f"MLP SVD energy={e}"
    elif subdir == "casp_slim":
        m = data.get("method", "")
        if m == "casp":
            return "Mixed 8/4-bit + LowRank QK"
        else:
            sp = data.get("target_sparsity", 0)
            rr = data.get("rank_ratio", 0)
            return f"SVD r={rr} + Prune {int(sp*100)}% + Quant"
    return ""

def family_sort_key(family):
    try:
        return FAMILY_ORDER.index(family)
    except ValueError:
        return len(FAMILY_ORDER)

def create_excel(records, output_path):
    wb = Workbook()

    # --- Sheet 1: All Results ---
    ws_all = wb.active
    ws_all.title = "All Results"

    columns = [
        "Model", "Family", "Method", "Configuration",
        "Params (M)", "Params After (M)", "GPU Mem Load (MB)", "Compression Ratio",
        "Accuracy", "Exact Match", "Contains", "Token F1", "BLEU", "ROUGE-L",
        "Avg Latency (s)", "Peak Memory (MB)", "Avg Memory (MB)",
        "Throughput (samples/s)", "Avg Power (W)", "GPU Util (%)",
        "Samples", "Evaluated", "Skipped", "Sparsity", "Quant Bits"
    ]

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Sort records
    method_order = [
        "Baseline (FP16)", "PTQ (BitsAndBytes)", "Magnitude Pruning", "Wanda",
        "SparseGPT", "AWQ (Sim. INT4)", "GPTQ (Sim. INT4)", "AWP (Prune+Quant)",
        "PACT (Prune+Merge)", "PALU (KV Compression)", "SVD-LLM",
        "CASP (Mixed Quant)", "SLIM (SVD+Prune+Quant)"
    ]

    def sort_key(r):
        fam = family_sort_key(r["Family"])
        try:
            meth = method_order.index(r["Method"])
        except ValueError:
            meth = len(method_order)
        return (fam, r["Model"], meth, r["Configuration"])

    records.sort(key=sort_key)

    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws_all.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = record.get(col_name, "")
            cell = ws_all.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            # Format numbers
            if isinstance(val, float):
                if col_name in ("Accuracy", "Exact Match", "Contains", "Token F1", "BLEU", "ROUGE-L", "Sparsity"):
                    cell.number_format = '0.0000'
                elif col_name in ("Compression Ratio",):
                    cell.number_format = '0.00'
                else:
                    cell.number_format = '0.0'

    # Auto-width
    for col_idx in range(1, len(columns) + 1):
        max_len = len(str(ws_all.cell(row=1, column=col_idx).value or ""))
        for row_idx in range(2, min(len(records) + 2, 20)):
            val = ws_all.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws_all.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 30)

    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(records)+1}"
    ws_all.freeze_panes = "A2"

    # --- Sheet 2: Summary by Model (accuracy comparison) ---
    ws_summary = wb.create_sheet("Accuracy Summary")

    # Build pivot: model -> method -> accuracy
    pivot = {}
    methods_seen = []
    for r in records:
        model = r["Model"]
        method = r["Method"]
        config = r["Configuration"]
        key = f"{method}\n({config})" if config else method
        acc = r["Accuracy"]
        if model not in pivot:
            pivot[model] = {}
        pivot[model][key] = acc
        if key not in methods_seen:
            methods_seen.append(key)

    # Write summary headers
    sum_cols = ["Model", "Family"] + methods_seen
    for col_idx, col_name in enumerate(sum_cols, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Get sorted unique models
    model_family = {}
    for r in records:
        model_family[r["Model"]] = r["Family"]

    sorted_models = sorted(pivot.keys(), key=lambda m: (family_sort_key(model_family.get(m, "")), m))

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_idx, model in enumerate(sorted_models, 2):
        ws_summary.cell(row=row_idx, column=1, value=model).border = thin_border
        ws_summary.cell(row=row_idx, column=2, value=model_family.get(model, "")).border = thin_border

        # Get baseline accuracy for this model
        baseline_acc = None
        for key in methods_seen:
            if "Baseline" in key:
                baseline_acc = pivot[model].get(key)
                break

        for col_idx, method_key in enumerate(methods_seen, 3):
            acc = pivot[model].get(method_key, None)
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=acc)
            cell.border = thin_border
            if isinstance(acc, (int, float)):
                cell.number_format = '0.0000'
                # Color: green if >= baseline, red if < baseline
                if baseline_acc is not None and isinstance(baseline_acc, (int, float)):
                    if acc >= baseline_acc:
                        cell.fill = green_fill
                    elif acc < baseline_acc * 0.9:  # >10% degradation
                        cell.fill = red_fill

    # Auto-width for summary
    for col_idx in range(1, len(sum_cols) + 1):
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = 22
    ws_summary.column_dimensions["A"].width = 40
    ws_summary.freeze_panes = "C2"

    # --- Sheet 3: Method Overview ---
    ws_methods = wb.create_sheet("Methods Overview")
    method_info = [
        ["Method", "Category", "Description", "Key Parameters"],
        ["Baseline (FP16)", "None", "Original model, no compression", "FP16 weights"],
        ["PTQ (BitsAndBytes)", "Quantization", "Post-training quantization via BitsAndBytes", "INT4 / INT8"],
        ["AWQ (Sim. INT4)", "Quantization", "Activation-aware weight quantization (simulated)", "4-bit, per-channel groups"],
        ["GPTQ (Sim. INT4)", "Quantization", "GPTQ quantization (simulated)", "4-bit, per-channel groups"],
        ["Magnitude Pruning", "Pruning", "L1 unstructured magnitude pruning", "Sparsity: 20%, 40%"],
        ["Wanda", "Pruning", "Pruning by weights and activations", "Sparsity: 20%, 40%"],
        ["SparseGPT", "Pruning", "One-shot pruning with Hessian-based updates", "Sparsity: 50%"],
        ["AWP (Prune+Quant)", "Combined", "Wanda pruning + simulated INT4 quantization", "Sparsity 50% + INT4"],
        ["PACT (Prune+Merge)", "Token Compression", "Progressive token pruning and merging", "Prune 30% + Merge 20%"],
        ["PALU (KV Compression)", "Low-Rank", "SVD-based KV cache compression", "Energy target 0.95"],
        ["SVD-LLM", "Low-Rank", "SVD decomposition of MLP layers", "Energy target 0.95"],
        ["CASP (Mixed Quant)", "Mixed Precision", "Channel-aware sensitivity-based mixed precision", "8/4-bit + LowRank QK"],
        ["SLIM (SVD+Prune+Quant)", "Combined", "SVD + pruning + quantization pipeline", "Rank 0.3 + Sparsity 50% + Quant"],
    ]

    for row_idx, row_data in enumerate(method_info, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_methods.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

    ws_methods.column_dimensions["A"].width = 30
    ws_methods.column_dimensions["B"].width = 18
    ws_methods.column_dimensions["C"].width = 55
    ws_methods.column_dimensions["D"].width = 35

    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")
    print(f"Total records: {len(records)}")
    print(f"Sheets: All Results, Accuracy Summary, Methods Overview")

if __name__ == "__main__":
    records = load_all_results()
    output_path = os.path.join(os.path.dirname(__file__), "..", "VLM_Compression_Benchmark_Results.xlsx")
    create_excel(records, output_path)
